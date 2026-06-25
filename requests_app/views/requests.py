# requests_app/views/requests.py
"""
Не-CRUD представления для заявок:
- назначение, выполнение, приостановка, возобновление, закрытие
- управление дополнительными исполнителями
- экспорт в Excel и настраиваемый отчёт

Используют сервисный слой, декораторы прав и логирование.
"""

from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
import logging
import csv
from decimal import Decimal

from users.models import UserRole
from ..models import (
    ServiceRequest, UsedMaterial, Material, RequestType,
    RequestHistory, RequestFile, RequestAssignee, RequestSettings
)
from ..forms import (
    ReportForm
)
from .permissions import can_edit_any_request, can_assign_request
from .decorators import (
    admin_required, manager_required, assign_required,
    can_view_request_required, can_edit_request_required,
    can_mark_completed_required, can_suspend_required,
    can_resume_required, can_close_request_required
)
from ..services import RequestService

logger = logging.getLogger(__name__)


# ============================================================
# Действия с заявками (изменение статуса, назначение)
# ============================================================


@login_required
@assign_required
def request_assign(request, pk):
    """
    Назначение основного исполнителя заявке.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if req.status not in ['new', 'in_progress']:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Неверный статус заявки'})
        messages.error(request, 'Неверный статус заявки.')
        return redirect('requests_app:request_detail', pk=pk)

    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')
        success, message = RequestService.assign_executor(req, assigned_to_id, request.user)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': success, 'message': message})
        if success:
            messages.success(request, message)
            logger.info(f"Заявка {req.request_number} назначена исполнителю {req.assigned_to}")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка назначения исполнителя для заявки {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=pk)
    return redirect('requests_app:request_detail', pk=pk)


@login_required
@can_mark_completed_required
def request_mark_completed(request, pk):
    """
    Отметка заявки как выполненной.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if request.method == 'POST':
        time_spent = request.POST.get('time_spent')
        success, message = RequestService.mark_completed(req, request.user, time_spent)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': success, 'message': message})
        if success:
            messages.success(request, message)
            logger.info(f"Заявка {req.request_number} отмечена выполненной")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка отметки выполнения заявки {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_mark_completed.html', {'req': req})


@login_required
@can_suspend_required
def request_suspend(request, pk):
    """
    Приостановка заявки с указанием причины.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('suspension_reason', '').strip()
        success, message = RequestService.suspend_request(req, request.user, reason)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': success, 'message': message})
        if success:
            messages.success(request, message)
            logger.info(f"Заявка {req.request_number} приостановлена. Причина: {reason}")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка приостановки заявки {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_suspend.html', {'req': req})


@login_required
@can_resume_required
def request_resume(request, pk):
    """
    Возобновление приостановленной или закрытой (для админа) заявки.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if request.method == 'POST':
        success, message = RequestService.resume_request(req, request.user)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': success, 'message': message})
        if success:
            messages.success(request, message)
            logger.info(f"Заявка {req.request_number} возобновлена")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка возобновления заявки {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_resume.html', {'req': req})


@login_required
@can_close_request_required
def request_close(request, pk):
    """
    Закрытие выполненной заявки со списанием материалов.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if request.method == 'POST':
        material_ids = request.POST.getlist('material_id[]')
        quantities = request.POST.getlist('material_quantity[]')
        units = request.POST.getlist('material_unit[]')
        prices = request.POST.getlist('material_price[]')
        materials_data = [
            {'material_id': mid, 'quantity': q, 'unit': u, 'price_per_unit': p}
            for mid, q, u, p in zip(material_ids, quantities, units, prices)
            if mid and q
        ]
        success, message = RequestService.close_request(req, request.user, materials_data)
        if success:
            messages.success(request, message)
            logger.info(f"Заявка {req.request_number} закрыта")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка закрытия заявки {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=req.pk)

    materials_qs = Material.objects.all().values('id', 'name', 'unit', 'default_price')
    materials_json = list(materials_qs)
    context = {
        'request_obj': req,
        'req': req,
        'materials': materials_qs,
        'materials_json': materials_json,
    }
    return render(request, 'requests_app/request_close.html', context)


# ============================================================
# Управление дополнительными исполнителями
# ============================================================


@login_required
@assign_required
def request_add_assignee(request, pk):
    """
    Добавление дополнительного исполнителя к заявке.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    if request.method == 'POST':
        assignee_id = request.POST.get('user_id')
        success, message = RequestService.add_assignee(req, request.user, assignee_id)
        if success:
            messages.success(request, message)
            logger.info(f"Добавлен исполнитель {assignee_id} к заявке {req.request_number}")
        else:
            messages.error(request, message)
            logger.warning(f"Ошибка добавления исполнителя к заявке {req.request_number}: {message}")
        return redirect('requests_app:request_detail', pk=pk)

    assigned_user_ids = req.assignees.values_list('user_id', flat=True)
    available_users = User.objects.filter(is_active=True).exclude(id__in=assigned_user_ids).exclude(id=req.assigned_to_id).order_by('username')
    return render(request, 'requests_app/add_assignee.html', {'request_obj': req, 'users': available_users})


@login_required
@assign_required
def request_remove_assignee(request, pk, user_id):
    """
    Удаление дополнительного исполнителя.
    """
    req = get_object_or_404(ServiceRequest, pk=pk)
    success, message = RequestService.remove_assignee(req, request.user, user_id)
    if success:
        messages.success(request, message)
        logger.info(f"Удалён исполнитель {user_id} из заявки {req.request_number}")
    else:
        messages.error(request, message)
        logger.warning(f"Ошибка удаления исполнителя из заявки {req.request_number}: {message}")
    return redirect('requests_app:request_detail', pk=pk)


# ============================================================
# Экспорт и отчёты
# ============================================================
@login_required
@manager_required
def custom_report(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'У вас нет доступа к отчётам.')
        return redirect('requests_app:request_list')

    # Сохранение выбранных колонок в сессию
    initial_columns = None
    if 'columns' in request.GET:
        request.session['report_columns'] = request.GET.getlist('columns')
    elif 'columns' not in request.GET:
        saved_columns = request.session.get('report_columns')
        if saved_columns:
            initial_columns = saved_columns

    form = ReportForm(request.GET or None, initial={'columns': initial_columns} if initial_columns else None)

    if role == UserRole.WORKER:
        qs = ServiceRequest.objects.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    else:
        qs = ServiceRequest.objects.all()

    if request.GET and form.is_valid():
        # сохраняем колонки в сессию
        request.session['report_columns'] = form.cleaned_data.get('columns', [])

        if form.cleaned_data.get('status'):
            qs = qs.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('priority'):
            qs = qs.filter(priority=form.cleaned_data['priority'])
        if form.cleaned_data.get('building'):
            qs = qs.filter(building=form.cleaned_data['building'])
        if form.cleaned_data.get('request_type'):
            qs = qs.filter(request_type=form.cleaned_data['request_type'])
        if form.cleaned_data.get('assigned_to'):
            qs = qs.filter(assigned_to=form.cleaned_data['assigned_to'])
        if form.cleaned_data.get('created_by'):
            qs = qs.filter(created_by=form.cleaned_data['created_by'])
        if form.cleaned_data.get('room_number'):
            qs = qs.filter(room_number__icontains=form.cleaned_data['room_number'])
        if form.cleaned_data.get('date_from'):
            qs = qs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            qs = qs.filter(created_at__date__lte=form.cleaned_data['date_to'])

    status_stats = qs.values('status').annotate(count=Count('id')).order_by('status')
    status_labels = []
    status_data = []
    status_display = dict(ServiceRequest.STATUS_CHOICES)
    for item in status_stats:
        status_labels.append(status_display.get(item['status'], item['status']))
        status_data.append(item['count'])

    columns = form.cleaned_data.get('columns') if form.is_valid() else []
    if not columns:
        columns = ['request_number', 'building', 'priority', 'status', 'created_by', 'assigned_to', 'created_at']

    field_map = {
        'request_number': '№ заявки',
        'building': 'Здание',
        'section': 'Часть здания',
        'room_number': 'Помещение',
        'request_type': 'Тип',
        'description': 'Описание',
        'priority': 'Приоритет',
        'status': 'Статус',
        'created_by': 'Создатель',
        'assigned_to': 'Ответственный',
        'planned_date': 'Плановая дата',
        'completed_date': 'Дата выполнения',
        'created_at': 'Дата создания',
        'comment': 'Комментарий',
    }

    # Экспорт в Excel
    if request.GET.get('export') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт по заявкам"
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_map.get(col, col))
            cell.font = openpyxl.styles.Font(bold=True)

        for row_idx, req in enumerate(qs.select_related('building', 'section', 'created_by', 'assigned_to', 'request_type'), 2):
            for col_idx, col in enumerate(columns, 1):
                value = ''
                if col == 'request_number':
                    value = req.request_number
                elif col == 'building':
                    value = str(req.building)
                elif col == 'section':
                    value = req.section.name if req.section else ''
                elif col == 'room_number':
                    value = req.room_number or ''
                elif col == 'request_type':
                    value = req.request_type.name if req.request_type else ''
                elif col == 'description':
                    value = req.description[:200] if req.description else ''
                elif col == 'priority':
                    value = req.get_priority_display()
                elif col == 'status':
                    value = req.get_status_display()
                elif col == 'created_by':
                    value = req.created_by.get_full_name() if req.created_by else (req.contact_name or 'Публичная')
                elif col == 'assigned_to':
                    value = req.assigned_to.get_full_name() if req.assigned_to else ''
                elif col == 'planned_date':
                    value = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
                elif col == 'completed_date':
                    value = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
                elif col == 'created_at':
                    value = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
                elif col == 'comment':
                    value = req.comment[:200] if req.comment else ''
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx in range(1, len(columns) + 1):
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="custom_report.xlsx"'
        try:
            wb.save(response)
            logger.info(f"Экспорт отчёта выполнен, записей: {qs.count()}")
            return response
        except Exception as e:
            logger.exception(f"Ошибка экспорта отчёта: {e}")
            messages.error(request, 'Ошибка при экспорте отчёта.')
            return redirect('requests_app:custom_report')

    # Экспорт в CSV
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="custom_report.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow([field_map.get(col, col) for col in columns])
        for req in qs.select_related('building', 'section', 'created_by', 'assigned_to', 'request_type'):
            row = []
            for col in columns:
                if col == 'request_number':
                    row.append(req.request_number)
                elif col == 'building':
                    row.append(str(req.building))
                elif col == 'section':
                    row.append(req.section.name if req.section else '')
                elif col == 'room_number':
                    row.append(req.room_number or '')
                elif col == 'request_type':
                    row.append(req.request_type.name if req.request_type else '')
                elif col == 'description':
                    row.append(req.description[:200] if req.description else '')
                elif col == 'priority':
                    row.append(req.get_priority_display())
                elif col == 'status':
                    row.append(req.get_status_display())
                elif col == 'created_by':
                    row.append(req.created_by.get_full_name() if req.created_by else (req.contact_name or 'Публичная'))
                elif col == 'assigned_to':
                    row.append(req.assigned_to.get_full_name() if req.assigned_to else '')
                elif col == 'planned_date':
                    row.append(req.planned_date.strftime('%d.%m.%Y') if req.planned_date else '')
                elif col == 'completed_date':
                    row.append(req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else '')
                elif col == 'created_at':
                    row.append(req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else '')
                elif col == 'comment':
                    row.append(req.comment[:200] if req.comment else '')
                else:
                    row.append('')
            writer.writerow(row)
        logger.info(f"Экспорт CSV выполнен, записей: {qs.count()}")
        return response

    # Отображение таблицы
    data = []
    for req in qs.select_related('building', 'section', 'created_by', 'assigned_to', 'request_type'):
        row = {}
        for col in columns:
            if col == 'request_number':
                row[col] = req.request_number
            elif col == 'building':
                row[col] = str(req.building)
            elif col == 'section':
                row[col] = req.section.name if req.section else ''
            elif col == 'room_number':
                row[col] = req.room_number or ''
            elif col == 'request_type':
                row[col] = req.request_type.name if req.request_type else ''
            elif col == 'description':
                row[col] = req.description[:100] if req.description else ''
            elif col == 'priority':
                row[col] = req.get_priority_display()
            elif col == 'status':
                row[col] = req.get_status_display()
            elif col == 'created_by':
                row[col] = req.created_by.get_full_name() if req.created_by else (req.contact_name or 'Публичная')
            elif col == 'assigned_to':
                row[col] = req.assigned_to.get_full_name() if req.assigned_to else ''
            elif col == 'planned_date':
                row[col] = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
            elif col == 'completed_date':
                row[col] = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
            elif col == 'created_at':
                row[col] = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
            elif col == 'comment':
                row[col] = req.comment[:200] if req.comment else ''
            else:
                row[col] = ''
        data.append(row)

    context = {
        'form': form,
        'data': data,
        'columns': columns,
        'column_labels': field_map,
        'field_map': field_map,
        'status_labels': status_labels,
        'status_data': status_data,
    }
    return render(request, 'requests_app/custom_report.html', context)

@login_required
@manager_required
def bulk_status_update(request):
    """
    Массовое изменение статуса выбранных заявок.
    Поддерживает AJAX-запросы с JSON-ответом.
    """
    if request.method != 'POST':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Метод не разрешён'}, status=405)
        return redirect('requests_app:request_list')

    ids = request.POST.getlist('request_ids[]') or request.POST.getlist('request_ids')
    if not ids:
        message = 'Не выбраны заявки.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    try:
        ids = [int(x) for x in ids if x and str(x).isdigit()]
    except ValueError:
        message = 'Некорректный ID заявки.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    if not ids:
        message = 'Не выбраны заявки.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    new_status = request.POST.get('new_status')
    if not new_status or new_status not in dict(ServiceRequest.STATUS_CHOICES):
        message = 'Некорректный статус.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    # Запрещаем массовый перевод в статусы, требующие дополнительных данных
    if new_status in ['closed', 'suspended']:
        message = 'Массовое изменение статуса на "Закрыта" или "Приостановлена" не поддерживается. Выполните эти действия индивидуально для каждой заявки.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        message = 'У вас нет прав на массовое изменение статуса.'
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': message})
        messages.error(request, message)
        return redirect('requests_app:request_list')

    # Для не-админов запрещаем закрытие, но оно уже отсечено выше

    qs = ServiceRequest.objects.filter(id__in=ids)
    if role != UserRole.ADMIN:
        qs = qs.exclude(status='closed')
    updated_count = qs.update(status=new_status)

    for req in qs:
        RequestHistory.objects.create(
            request=req,
            user=user,
            action=f'Статус изменён на "{dict(ServiceRequest.STATUS_CHOICES).get(new_status, new_status)}" (массовое обновление)'
        )

    message = f'Статус обновлён для {updated_count} заявок.'
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': message, 'updated': updated_count})
    messages.success(request, message)
    return redirect('requests_app:request_list')