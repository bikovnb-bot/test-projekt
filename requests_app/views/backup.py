# requests_app/views/backup.py
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.models import User

from buildings.models import Building, BuildingSection
from users.models import UserRole
from ..models import ServiceRequest, UsedMaterial, RequestType, RequestFile, RequestAssignee, RequestHistory
from .decorators import manager_required
from ..utils import parse_date, parse_datetime

logger = logging.getLogger(__name__)


@login_required
@manager_required
def export_requests_full_backup(request):
    """
    Экспорт всех данных заявок в Excel-бэкап.
    """
    wb = openpyxl.Workbook()
    ws_requests = wb.active
    ws_requests.title = "Заявки"
    request_headers = [
        'request_number', 'building_address', 'section_name', 'room_number', 'request_type_name',
        'description', 'priority', 'status', 'created_by_username', 'assigned_to_username',
        'planned_date', 'completed_date', 'comment', 'created_at', 'updated_at',
        'track_time', 'time_spent', 'suspension_reason', 'contact_name', 'contact_phone', 'ip_address'
    ]
    ws_requests.append(request_headers)

    qs = ServiceRequest.objects.select_related('building', 'section', 'request_type', 'created_by', 'assigned_to').all()
    for req in qs.iterator(chunk_size=500):
        ws_requests.append([
            req.request_number,
            req.building.address if req.building else '',
            req.section.name if req.section else '',
            req.room_number or '',
            req.request_type.name if req.request_type else '',
            req.description or '',
            req.priority,
            req.status,
            req.created_by.username if req.created_by else (req.contact_name or ''),
            req.assigned_to.username if req.assigned_to else '',
            req.planned_date.strftime('%Y-%m-%d') if req.planned_date else '',
            req.completed_date.strftime('%Y-%m-%d %H:%M:%S') if req.completed_date else '',
            req.comment or '',
            req.created_at.strftime('%Y-%m-%d %H:%M:%S') if req.created_at else '',
            req.updated_at.strftime('%Y-%m-%d %H:%M:%S') if req.updated_at else '',
            req.track_time,
            req.time_spent or '',
            req.suspension_reason or '',
            req.contact_name or '',
            req.contact_phone or '',
            req.ip_address or '',
        ])

    ws_materials = wb.create_sheet("Материалы")
    material_headers = ['request_number', 'material_name', 'quantity', 'unit', 'price_per_unit', 'total_price']
    ws_materials.append(material_headers)
    used_materials = UsedMaterial.objects.select_related('request').all()
    for um in used_materials.iterator(chunk_size=500):
        ws_materials.append([
            um.request.request_number,
            um.name,
            float(um.quantity),
            um.unit,
            float(um.price_per_unit),
            float(um.total_price) if um.total_price else '',
        ])

    ws_files = wb.create_sheet("Файлы")
    file_headers = ['request_number', 'file_name', 'description', 'uploaded_by_username', 'uploaded_at']
    ws_files.append(file_headers)
    files = RequestFile.objects.select_related('request', 'uploaded_by').all()
    for f in files.iterator(chunk_size=500):
        ws_files.append([
            f.request.request_number,
            f.get_file_name(),
            f.description or '',
            f.uploaded_by.username if f.uploaded_by else '',
            f.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if f.uploaded_at else '',
        ])

    ws_assignees = wb.create_sheet("Исполнители")
    assignee_headers = ['request_number', 'user_username']
    ws_assignees.append(assignee_headers)
    assignees = RequestAssignee.objects.select_related('request', 'user').all()
    for a in assignees.iterator(chunk_size=500):
        ws_assignees.append([
            a.request.request_number,
            a.user.username,
        ])

    ws_history = wb.create_sheet("История")
    history_headers = ['request_number', 'user_username', 'action', 'old_value', 'new_value', 'created_at']
    ws_history.append(history_headers)
    history = RequestHistory.objects.select_related('request', 'user').all()
    for h in history.iterator(chunk_size=500):
        ws_history.append([
            h.request.request_number,
            h.user.username if h.user else '',
            h.action,
            h.old_value or '',
            h.new_value or '',
            h.created_at.strftime('%Y-%m-%d %H:%M:%S') if h.created_at else '',
        ])

    for ws in [ws_requests, ws_materials, ws_files, ws_assignees, ws_history]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    filename = f"requests_full_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@manager_required
def import_requests_full_backup(request):
    """
    Импорт данных из Excel-бэкапа заявок.
    """
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        if not backup_file.name.endswith('.xlsx'):
            messages.error(request, 'Поддерживаются только файлы .xlsx')
            return redirect('requests_app:request_list')

        try:
            wb = openpyxl.load_workbook(backup_file, data_only=True)
            required_sheets = ['Заявки', 'Материалы', 'Файлы', 'Исполнители', 'История']
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    messages.error(request, f'В файле отсутствует лист "{sheet}"')
                    return redirect('requests_app:request_list')

            ws_requests = wb['Заявки']
            request_headers = [cell.value for cell in ws_requests[1]]
            req_col_idx = {h: i for i, h in enumerate(request_headers) if h}
            required_req_fields = ['request_number', 'building_address', 'request_type_name']
            for f in required_req_fields:
                if f not in req_col_idx:
                    messages.error(request, f'В листе "Заявки" отсутствует колонка "{f}"')
                    return redirect('requests_app:request_list')

            created_count = 0
            updated_count = 0
            errors = []
            request_numbers_in_backup = []
            for row in ws_requests.iter_rows(min_row=2, values_only=True):
                if row and row[req_col_idx['request_number']]:
                    request_numbers_in_backup.append(str(row[req_col_idx['request_number']]).strip())

            for row_idx, row in enumerate(ws_requests.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(c is None for c in row):
                    continue
                contract_number = str(row[req_col_idx['request_number']]).strip() if row[req_col_idx['request_number']] else None
                if not contract_number:
                    errors.append(f'Строка {row_idx}: отсутствует номер заявки')
                    continue

                building_address = str(row[req_col_idx['building_address']]).strip() if row[req_col_idx['building_address']] else None
                building = None
                if building_address:
                    building = Building.objects.filter(address__iexact=building_address).first()
                    if not building:
                        errors.append(f'Строка {row_idx}: здание "{building_address}" не найдено')
                        continue

                section_name = str(row[req_col_idx.get('section_name', 2)]).strip() if req_col_idx.get('section_name') and row[req_col_idx.get('section_name')] else None
                section = None
                if section_name and building:
                    section = BuildingSection.objects.filter(building=building, name__iexact=section_name).first()
                    if not section:
                        errors.append(f'Строка {row_idx}: часть здания "{section_name}" не найдена для здания {building.address}')

                request_type_name = str(row[req_col_idx['request_type_name']]).strip() if row[req_col_idx['request_type_name']] else None
                request_type = None
                if request_type_name:
                    request_type = RequestType.objects.filter(name__iexact=request_type_name).first()
                    if not request_type:
                        errors.append(f'Строка {row_idx}: тип заявки "{request_type_name}" не найден')
                        continue

                created_by_username = str(row[req_col_idx.get('created_by_username', 8)]).strip() if req_col_idx.get('created_by_username') and row[req_col_idx.get('created_by_username')] else None
                created_by = None
                if created_by_username:
                    created_by = User.objects.filter(username=created_by_username).first()

                assigned_to_username = str(row[req_col_idx.get('assigned_to_username', 9)]).strip() if req_col_idx.get('assigned_to_username') and row[req_col_idx.get('assigned_to_username')] else None
                assigned_to = None
                if assigned_to_username:
                    assigned_to = User.objects.filter(username=assigned_to_username).first()

                planned_date = parse_date(row[req_col_idx.get('planned_date', 10)]) if req_col_idx.get('planned_date') and row[req_col_idx.get('planned_date')] else None
                completed_date = parse_datetime(row[req_col_idx.get('completed_date', 11)]) if req_col_idx.get('completed_date') and row[req_col_idx.get('completed_date')] else None
                created_at = parse_datetime(row[req_col_idx.get('created_at', 13)]) if req_col_idx.get('created_at') and row[req_col_idx.get('created_at')] else None
                updated_at = parse_datetime(row[req_col_idx.get('updated_at', 14)]) if req_col_idx.get('updated_at') and row[req_col_idx.get('updated_at')] else None

                defaults = {
                    'building': building,
                    'section': section,
                    'room_number': str(row[req_col_idx.get('room_number', 3)]).strip() if req_col_idx.get('room_number') and row[req_col_idx.get('room_number')] else '',
                    'request_type': request_type,
                    'description': str(row[req_col_idx.get('description', 5)]).strip() if req_col_idx.get('description') and row[req_col_idx.get('description')] else '',
                    'priority': str(row[req_col_idx.get('priority', 6)]).strip() if req_col_idx.get('priority') and row[req_col_idx.get('priority')] else 'medium',
                    'status': str(row[req_col_idx.get('status', 7)]).strip() if req_col_idx.get('status') and row[req_col_idx.get('status')] else 'new',
                    'created_by': created_by,
                    'assigned_to': assigned_to,
                    'planned_date': planned_date,
                    'completed_date': completed_date,
                    'comment': str(row[req_col_idx.get('comment', 12)]).strip() if req_col_idx.get('comment') and row[req_col_idx.get('comment')] else '',
                    'created_at': created_at,
                    'updated_at': updated_at,
                    'track_time': bool(row[req_col_idx.get('track_time', 15)]) if req_col_idx.get('track_time') and row[req_col_idx.get('track_time')] else False,
                    'time_spent': int(row[req_col_idx.get('time_spent', 16)]) if req_col_idx.get('time_spent') and row[req_col_idx.get('time_spent')] else None,
                    'suspension_reason': str(row[req_col_idx.get('suspension_reason', 17)]).strip() if req_col_idx.get('suspension_reason') and row[req_col_idx.get('suspension_reason')] else '',
                    'contact_name': str(row[req_col_idx.get('contact_name', 18)]).strip() if req_col_idx.get('contact_name') and row[req_col_idx.get('contact_name')] else '',
                    'contact_phone': str(row[req_col_idx.get('contact_phone', 19)]).strip() if req_col_idx.get('contact_phone') and row[req_col_idx.get('contact_phone')] else '',
                    'ip_address': str(row[req_col_idx.get('ip_address', 20)]).strip() if req_col_idx.get('ip_address') and row[req_col_idx.get('ip_address')] else '',
                }
                try:
                    req_obj, created = ServiceRequest.objects.update_or_create(
                        request_number=contract_number,
                        defaults=defaults
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append(f'Строка {row_idx} (заявка {contract_number}): {str(e)}')
                    logger.error(f"Ошибка импорта заявки {contract_number}: {e}")

            # Импорт материалов
            if 'Материалы' in wb.sheetnames:
                ws_materials = wb['Материалы']
                material_headers = [cell.value for cell in ws_materials[1]]
                m_col_idx = {h: i for i, h in enumerate(material_headers) if h}
                if 'request_number' in m_col_idx and 'material_name' in m_col_idx:
                    UsedMaterial.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_materials.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[m_col_idx['request_number']]).strip() if row[m_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Материалы строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Материалы строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        material_name = str(row[m_col_idx['material_name']]).strip() if row[m_col_idx['material_name']] else None
                        if not material_name:
                            errors.append(f'Материалы строка {row_idx}: нет названия материала')
                            continue
                        quantity = Decimal(str(row[m_col_idx['quantity']]).replace(',', '.')) if row[m_col_idx['quantity']] else Decimal('0')
                        unit = str(row[m_col_idx['unit']]).strip() if row[m_col_idx['unit']] else ''
                        price_per_unit = Decimal(str(row[m_col_idx['price_per_unit']]).replace(',', '.')) if row[m_col_idx['price_per_unit']] else Decimal('0')
                        total_price = Decimal(str(row[m_col_idx['total_price']]).replace(',', '.')) if m_col_idx.get('total_price') and row[m_col_idx['total_price']] else quantity * price_per_unit
                        try:
                            UsedMaterial.objects.create(
                                request=req_obj,
                                material=None,
                                name=material_name,
                                quantity=quantity,
                                unit=unit,
                                price_per_unit=price_per_unit,
                                total_price=total_price
                            )
                        except Exception as e:
                            errors.append(f'Материалы строка {row_idx}: {str(e)}')
                            logger.error(f"Ошибка импорта материала {material_name}: {e}")
                else:
                    errors.append('Лист "Материалы" не имеет обязательных колонок request_number, material_name')
                    logger.warning("Лист Материалы не имеет обязательных колонок")

            # Импорт файлов
            if 'Файлы' in wb.sheetnames:
                ws_files = wb['Файлы']
                file_headers = [cell.value for cell in ws_files[1]]
                f_col_idx = {h: i for i, h in enumerate(file_headers) if h}
                if 'request_number' in f_col_idx and 'file_name' in f_col_idx:
                    RequestFile.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_files.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[f_col_idx['request_number']]).strip() if row[f_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Файлы строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Файлы строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        file_name = str(row[f_col_idx['file_name']]).strip()
                        description = str(row[f_col_idx.get('description', 2)]).strip() if f_col_idx.get('description') and row[f_col_idx.get('description')] else ''
                        uploaded_by_username = str(row[f_col_idx.get('uploaded_by_username', 3)]).strip() if f_col_idx.get('uploaded_by_username') and row[f_col_idx.get('uploaded_by_username')] else None
                        uploaded_by = None
                        if uploaded_by_username:
                            uploaded_by = User.objects.filter(username=uploaded_by_username).first()
                        uploaded_at = parse_datetime(row[f_col_idx.get('uploaded_at', 4)]) if f_col_idx.get('uploaded_at') and row[f_col_idx.get('uploaded_at')] else None
                        try:
                            RequestFile.objects.create(
                                request=req_obj,
                                file=None,
                                uploaded_by=uploaded_by,
                                uploaded_at=uploaded_at or timezone.now(),
                                description=description
                            )
                        except Exception as e:
                            errors.append(f'Файлы строка {row_idx}: {str(e)}')
                            logger.error(f"Ошибка импорта файла {file_name}: {e}")
                else:
                    errors.append('Лист "Файлы" не имеет обязательных колонок request_number, file_name')
                    logger.warning("Лист Файлы не имеет обязательных колонок")

            # Импорт исполнителей
            if 'Исполнители' in wb.sheetnames:
                ws_assignees = wb['Исполнители']
                assignee_headers = [cell.value for cell in ws_assignees[1]]
                a_col_idx = {h: i for i, h in enumerate(assignee_headers) if h}
                if 'request_number' in a_col_idx and 'user_username' in a_col_idx:
                    RequestAssignee.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_assignees.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[a_col_idx['request_number']]).strip() if row[a_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Исполнители строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Исполнители строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        username = str(row[a_col_idx['user_username']]).strip() if row[a_col_idx['user_username']] else None
                        if not username:
                            errors.append(f'Исполнители строка {row_idx}: нет имени пользователя')
                            continue
                        user_obj = User.objects.filter(username=username).first()
                        if not user_obj:
                            errors.append(f'Исполнители строка {row_idx}: пользователь {username} не найден')
                            continue
                        try:
                            RequestAssignee.objects.get_or_create(request=req_obj, user=user_obj)
                        except Exception as e:
                            errors.append(f'Исполнители строка {row_idx}: {str(e)}')
                            logger.error(f"Ошибка импорта исполнителя {username}: {e}")
                else:
                    errors.append('Лист "Исполнители" не имеет обязательных колонок request_number, user_username')
                    logger.warning("Лист Исполнители не имеет обязательных колонок")

            # Импорт истории
            if 'История' in wb.sheetnames:
                ws_history = wb['История']
                history_headers = [cell.value for cell in ws_history[1]]
                h_col_idx = {h: i for i, h in enumerate(history_headers) if h}
                if 'request_number' in h_col_idx and 'action' in h_col_idx:
                    RequestHistory.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_history.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[h_col_idx['request_number']]).strip() if row[h_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'История строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'История строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        action = str(row[h_col_idx['action']]).strip() if row[h_col_idx['action']] else ''
                        if not action:
                            errors.append(f'История строка {row_idx}: нет действия')
                            continue
                        user_username = str(row[h_col_idx.get('user_username', 1)]).strip() if h_col_idx.get('user_username') and row[h_col_idx.get('user_username')] else None
                        user_obj = None
                        if user_username:
                            user_obj = User.objects.filter(username=user_username).first()
                        old_value = str(row[h_col_idx.get('old_value', 2)]).strip() if h_col_idx.get('old_value') and row[h_col_idx.get('old_value')] else ''
                        new_value = str(row[h_col_idx.get('new_value', 3)]).strip() if h_col_idx.get('new_value') and row[h_col_idx.get('new_value')] else ''
                        created_at = parse_datetime(row[h_col_idx.get('created_at', 4)]) if h_col_idx.get('created_at') and row[h_col_idx.get('created_at')] else None
                        try:
                            RequestHistory.objects.create(
                                request=req_obj,
                                user=user_obj,
                                action=action,
                                old_value=old_value,
                                new_value=new_value,
                                created_at=created_at or timezone.now()
                            )
                        except Exception as e:
                            errors.append(f'История строка {row_idx}: {str(e)}')
                            logger.error(f"Ошибка импорта записи истории для {request_number}: {e}")
                else:
                    errors.append('Лист "История" не имеет обязательных колонок request_number, action')
                    logger.warning("Лист История не имеет обязательных колонок")

            msg = f'Импорт завершён. Заявки: создано {created_count}, обновлено {updated_count}.'
            if errors:
                msg += f' Ошибки ({len(errors)}): {", ".join(errors[:10])}'
                messages.warning(request, msg)
            else:
                messages.success(request, msg)

        except Exception as e:
            logger.exception("Ошибка импорта бэкапа заявок")
            messages.error(request, f'Ошибка обработки файла: {str(e)}')
        return redirect('requests_app:request_list')

    return render(request, 'requests_app/import_requests_backup.html')