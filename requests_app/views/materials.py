# requests_app/views/materials.py
"""
Представления для управления материалами на складе.
Используют сервисный слой для работы с данными.
"""

import logging
import openpyxl
from decimal import Decimal
from datetime import datetime, date

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper
from django.db import transaction
from django.utils import timezone
from django import forms

from users.models import UserRole
from ..models import Material, UsedMaterial, MaterialTransaction
from ..forms import MaterialForm, ImportMaterialsForm, MaterialAdjustForm
from .decorators import manager_required
from ..services import MaterialService, NotificationService

logger = logging.getLogger(__name__)


# ===== Существующие представления =====

@login_required
def material_stock(request):
    """
    Отображает список материалов на складе с поиском и пагинацией.
    """
    search_query = request.GET.get('search', '').strip()
    materials_qs = MaterialService.get_material_stock()
    if search_query:
        materials_qs = materials_qs.filter(name__icontains=search_query)
    paginator = Paginator(materials_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    user_role = request.user.profile.role if hasattr(request.user, 'profile') else None
    context = {
        'materials': page_obj,
        'search': search_query,
        'user_role': user_role,
    }
    return render(request, 'requests_app/material_stock.html', context)


@login_required
@manager_required
def material_add(request):
    """
    Добавление нового материала.
    """
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            try:
                material = MaterialService.create_material(form.cleaned_data)
                messages.success(request, f'Материал "{material.name}" добавлен.')
                logger.info(f"Добавлен материал: {material.name} (id={material.id})")
            except Exception as e:
                logger.exception(f"Ошибка добавления материала: {e}")
                messages.error(request, 'Ошибка при добавлении материала.')
            return redirect('requests_app:material_stock')
    else:
        form = MaterialForm()
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Добавить материал'})


@login_required
@manager_required
def material_edit(request, pk):
    """
    Редактирование материала.
    """
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            try:
                material = MaterialService.update_material(material, form.cleaned_data)
                messages.success(request, f'Материал "{material.name}" успешно обновлён.')
                logger.info(f"Обновлён материал: {material.name} (id={material.id})")
            except Exception as e:
                logger.exception(f"Ошибка обновления материала {material.id}: {e}")
                messages.error(request, 'Ошибка при обновлении материала.')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме.')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Редактировать материал'})


@login_required
@manager_required
def material_delete(request, pk):
    """
    Удаление материала через POST (обычная форма).
    """
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        name = material.name
        try:
            MaterialService.delete_material(material)
            messages.success(request, f'Материал "{name}" удалён.')
            logger.info(f"Удалён материал: {name} (id={material.id})")
        except Exception as e:
            logger.exception(f"Ошибка удаления материала {material.id}: {e}")
            messages.error(request, 'Ошибка при удалении материала.')
        return redirect('requests_app:material_stock')
    return render(request, 'requests_app/material_confirm_delete.html', {'material': material})


@login_required
@manager_required
def material_delete_ajax(request, pk):
    """
    AJAX-удаление материала.
    """
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        try:
            MaterialService.delete_material(material)
            logger.info(f"AJAX-удалён материал: {material.name} (id={material.id})")
            return JsonResponse({'success': True})
        except Exception as e:
            logger.exception(f"Ошибка AJAX-удаления материала {material.id}: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)


@login_required
def material_stock_export(request):
    """
    Экспорт списка материалов в Excel.
    """
    search = request.GET.get('search', '').strip()
    materials_qs = Material.objects.all()
    if search:
        materials_qs = materials_qs.filter(name__icontains=search)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Материалы"
    headers = ['Наименование', 'Единица измерения', 'Количество', 'Цена за ед.', 'Минимальный остаток']
    ws.append(headers)
    
    for item in materials_qs:
        ws.append([
            item.name,
            item.unit,
            float(item.quantity_in_stock),
            float(item.default_price),
            float(item.min_stock)
        ])
    
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="material_stock_export.xlsx"'
    wb.save(response)
    logger.info(f"Экспорт материалов выполнен, записей: {materials_qs.count()}")
    return response


@login_required
@manager_required
def import_materials(request):
    """
    Импорт материалов из Excel-файла.
    """
    if request.method == 'POST' and request.FILES.get('excel_file'):
        form = ImportMaterialsForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                created = 0
                updated = 0
                with transaction.atomic():
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 5:
                            continue
                        name = row[0] if row[0] else None
                        unit = row[1] if len(row) > 1 and row[1] else None
                        default_price = row[2] if len(row) > 2 and row[2] is not None else 0
                        quantity_in_stock = row[3] if len(row) > 3 and row[3] is not None else 0
                        min_stock = row[4] if len(row) > 4 and row[4] is not None else 0
                        if name and unit:
                            try:
                                default_price = float(str(default_price).replace(',', '.'))
                            except (ValueError, TypeError):
                                logger.warning(f"Неверный формат цены для материала {name}: {default_price}")
                                default_price = 0.0
                            try:
                                quantity_in_stock = float(str(quantity_in_stock).replace(',', '.'))
                            except (ValueError, TypeError):
                                logger.warning(f"Неверный формат количества для материала {name}: {quantity_in_stock}")
                                quantity_in_stock = 0.0
                            try:
                                min_stock = float(str(min_stock).replace(',', '.'))
                            except (ValueError, TypeError):
                                logger.warning(f"Неверный формат минимального остатка для материала {name}: {min_stock}")
                                min_stock = 0.0
                            material, is_created = Material.objects.update_or_create(
                                name=name,
                                defaults={
                                    'unit': unit,
                                    'default_price': default_price,
                                    'quantity_in_stock': quantity_in_stock,
                                    'min_stock': min_stock
                                }
                            )
                            if is_created:
                                created += 1
                            else:
                                updated += 1
                messages.success(request, f'Импортировано: добавлено {created}, обновлено {updated}.')
                logger.info(f"Импорт материалов: добавлено {created}, обновлено {updated}")
            except openpyxl.utils.exceptions.InvalidFileException as e:
                logger.error(f"Неверный формат Excel-файла: {e}")
                messages.error(request, 'Неверный формат файла. Загрузите корректный .xlsx.')
            except Exception as e:
                logger.exception(f"Ошибка импорта материалов: {e}")
                messages.error(request, f'Ошибка обработки файла: {str(e)}')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме. Проверьте файл.')
    else:
        form = ImportMaterialsForm()
    return render(request, 'requests_app/import_materials.html', {'form': form})


@login_required
def download_materials_template(request):
    """
    Скачивание шаблона Excel для импорта материалов.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Материалы"
    headers = ['name', 'unit', 'default_price', 'quantity_in_stock', 'min_stock']
    ws.append(headers)
    ws.append(['Краска', 'л', 350.00, 100, 10])
    ws.append(['Лампа светодиодная', 'шт', 450.00, 50, 5])
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="materials_import_template.xlsx"'
    wb.save(response)
    logger.info("Скачан шаблон импорта материалов")
    return response


@login_required
@manager_required
def material_history(request, pk):
    """
    Просмотр истории транзакций материала.
    """
    material = get_object_or_404(Material, pk=pk)
    transactions = MaterialService.get_transactions(material)
    context = {
        'material': material,
        'transactions': transactions,
    }
    return render(request, 'requests_app/material_history.html', context)


@login_required
@manager_required
def material_adjust(request, pk):
    """
    Корректировка остатка материала (приход/списание).
    """
    material = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        form = MaterialAdjustForm(request.POST)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            transaction_type = form.cleaned_data['transaction_type']
            comment = form.cleaned_data.get('comment', '')
            try:
                MaterialService.adjust_stock(material, quantity, transaction_type, comment)
                messages.success(request, f'Остаток материала "{material.name}" скорректирован.')
                logger.info(f"Корректировка материала {material.name}: {transaction_type} {quantity} (комментарий: {comment})")
                return redirect('requests_app:material_stock')
            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.exception(f"Ошибка корректировки материала {material.id}: {e}")
                messages.error(request, 'Ошибка при корректировке остатка.')
    else:
        form = MaterialAdjustForm()
    return render(request, 'requests_app/material_adjust.html', {'form': form, 'material': material})


# ===== НОВОЕ ПРЕДСТАВЛЕНИЕ: ОТЧЁТ ПО РАСХОДУ МАТЕРИАЛОВ =====

class ConsumptionReportForm(forms.Form):
    """Форма для выбора периода в отчёте по расходу материалов."""
    date_from = forms.DateField(
        required=True,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
        label="Дата начала"
    )
    date_to = forms.DateField(
        required=True,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
        label="Дата окончания"
    )
    material = forms.ModelChoiceField(
        queryset=Material.objects.all().order_by('name'),
        required=False,
        widget=forms.Select(attrs={'class': 'form-select'}),
        label="Материал (все)"
    )

    def clean(self):
        cleaned_data = super().clean()
        date_from = cleaned_data.get('date_from')
        date_to = cleaned_data.get('date_to')
        if date_from and date_to and date_from > date_to:
            raise forms.ValidationError('Дата начала не может быть позже даты окончания.')
        return cleaned_data


@login_required
@manager_required
def material_consumption_report(request):
    """
    Отчёт по расходу материалов за выбранный период.
    Выводит таблицу с количеством и суммой по каждому материалу.
    Поддерживает экспорт в Excel.
    """
    form = ConsumptionReportForm(request.GET or None)
    report_data = []
    total_quantity = 0
    total_sum = 0
    date_from = None
    date_to = None
    selected_material = None

    if form.is_valid():
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']
        selected_material = form.cleaned_data.get('material')

        # Базовый queryset UsedMaterial
        qs = UsedMaterial.objects.filter(
            request__created_at__date__gte=date_from,
            request__created_at__date__lte=date_to
        )

        if selected_material:
            qs = qs.filter(material=selected_material)

        # Группировка по материалу
        report_data = (
            qs.values('material__id', 'material__name', 'material__unit')
            .annotate(
                total_quantity=Sum('quantity'),
                total_sum=Sum('total_price')
            )
            .order_by('material__name')
        )

        total_quantity = sum(item['total_quantity'] for item in report_data)
        total_sum = sum(item['total_sum'] for item in report_data)

    # Экспорт в Excel
    if request.GET.get('export') == '1' and form.is_valid():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расход материалов"
        headers = ['№', 'Наименование', 'Единица измерения', 'Количество', 'Сумма (₽)']
        ws.append(headers)

        for idx, item in enumerate(report_data, start=1):
            ws.append([
                idx,
                item['material__name'],
                item['material__unit'],
                float(item['total_quantity']),
                float(item['total_sum'])
            ])

        # Итоговая строка
        ws.append([])
        ws.append(['', 'ИТОГО', '', float(total_quantity), float(total_sum)])

        # Автоширина колонок
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"consumption_report_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        logger.info(f"Экспорт отчёта по расходу материалов выполнен за период {date_from} - {date_to}")
        return response

    context = {
        'form': form,
        'report_data': report_data,
        'total_quantity': total_quantity,
        'total_sum': total_sum,
        'date_from': date_from,
        'date_to': date_to,
        'selected_material': selected_material,
    }
    return render(request, 'requests_app/material_consumption_report.html', context)