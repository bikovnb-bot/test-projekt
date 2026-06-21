# requests_app/views.py
# Полный файл с удалёнными неиспользуемыми блоками дашборда (виджеты, настройки)
# и добавленной функцией get_dashboard_context, возвращающей user_stats (все исполнители)

import os
import json
import random
from decimal import Decimal
from datetime import datetime, timedelta, time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.urls import reverse

from buildings.models import Building, BuildingSection
from users.models import UserRole
from .models import (
    ServiceRequest, UsedMaterial, Material, RequestType,
    RequestHistory, RequestFile, RequestAssignee, RequestSettings
)
from .forms import (
    ServiceRequestForm, UsedMaterialFormSet, ReportForm,
    ImportMaterialsForm, MaterialForm, PublicRequestForm
)
from .utils import send_telegram_notification, rate_limit
from .translator import translate_to_russian


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def generate_new_captcha(lang='ru'):
    operators = ['+', '-', '*']
    op = random.choice(operators)
    if op == '+':
        a = random.randint(1, 20)
        b = random.randint(1, 20)
    elif op == '-':
        a = random.randint(5, 20)
        b = random.randint(1, a)
    else:
        a = random.randint(1, 10)
        b = random.randint(1, 10)
    return {
        'captcha_num1': a,
        'captcha_num2': b,
        'captcha_operator': op,
    }

def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except:
            pass
    if isinstance(value, str):
        for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None

def parse_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except:
            pass
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%Y-%m-%d'):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


# ==================== ФУНКЦИЯ ДЛЯ ПОЛУЧЕНИЯ КОНТЕКСТА ДАШБОРДА ====================

def get_dashboard_context(year=None, month=None):
    """
    Возвращает контекст для дашборда заявок.
    Параметры:
        year  - год (по умолчанию текущий)
        month - месяц (по умолчанию текущий)
    Используется как в старом представлении request_dashboard, так и в общем дашборде.
    """
    import calendar
    from datetime import date
    from django.db.models import Q, Count

    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    qs = ServiceRequest.objects.filter(
        created_at__year=year,
        created_at__month=month
    )

    total_requests = qs.count()
    completed_closed = qs.filter(status__in=['completed', 'closed']).count()
    in_progress = qs.filter(status='in_progress').count()
    overdue = qs.filter(planned_date__lt=date.today(), status__in=['new', 'in_progress', 'suspended']).count()
    completion_rate = round(completed_closed / total_requests * 100, 1) if total_requests else 0

    num_days = calendar.monthrange(year, month)[1]
    days_list = list(range(1, num_days + 1))
    day_created = [qs.filter(created_at__day=d).count() for d in days_list]
    day_completed = [qs.filter(completed_date__day=d, status__in=['completed', 'closed']).count() for d in days_list]

    year_requests = ServiceRequest.objects.filter(created_at__year=year)
    monthly_created = [year_requests.filter(created_at__month=m).count() for m in range(1, 13)]
    month_labels = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']

    status_counts = qs.values('status').annotate(cnt=Count('id'))
    status_dict = dict(ServiceRequest.STATUS_CHOICES)
    status_labels = [status_dict.get(s['status'], s['status']) for s in status_counts]
    status_data = [s['cnt'] for s in status_counts]

    type_counts = qs.values('request_type__name').annotate(cnt=Count('id')).order_by('-cnt')[:5]
    type_labels = [t['request_type__name'] or 'Без типа' for t in type_counts]
    type_data = [t['cnt'] for t in type_counts]

    priority_counts = qs.values('priority').annotate(cnt=Count('id'))
    priority_dict = dict(ServiceRequest.PRIORITY_CHOICES)
    priority_labels = [priority_dict.get(p['priority'], p['priority']) for p in priority_counts]
    priority_data = [p['cnt'] for p in priority_counts]

    executor_counts = {}
    for req in qs.select_related('assigned_to').prefetch_related('assignees__user'):
        if req.assigned_to:
            executor_counts[req.assigned_to] = executor_counts.get(req.assigned_to, 0) + 1
        for assignee in req.assignees.all():
            user_obj = assignee.user
            executor_counts[user_obj] = executor_counts.get(user_obj, 0) + 1

    top_executors_raw = sorted(
        [{'user': u, 'count': c} for u, c in executor_counts.items()],
        key=lambda x: x['count'], reverse=True
    )[:5]
    top_executors = [{'name': u.get_full_name() or u.username, 'count': c} for u, c in [(item['user'], item['count']) for item in top_executors_raw]]

    top_users = [item['user'] for item in top_executors_raw]
    assignee_daily = []
    for user_obj in top_users:
        daily_counts = []
        for day in days_list:
            day_requests = qs.filter(created_at__day=day)
            count = day_requests.filter(assigned_to=user_obj).count()
            count += day_requests.filter(assignees__user=user_obj).count()
            daily_counts.append(count)
        assignee_daily.append({
            'label': user_obj.get_full_name() or user_obj.username,
            'data': daily_counts
        })

    assignee_list = [{'name': u.get_full_name() or u.username, 'total': c} for u, c in executor_counts.items()]
    assignee_list.sort(key=lambda x: x['total'], reverse=True)
    assignee_list = assignee_list[:5]

    completed_reqs = qs.filter(status__in=['completed', 'closed'], completed_date__isnull=False)
    avg_hours = 0
    if completed_reqs.exists():
        total_seconds = sum((req.completed_date - req.created_at).total_seconds() for req in completed_reqs)
        avg_hours = total_seconds / len(completed_reqs) / 3600

    # --- Эффективность сотрудников (все, кому назначались заявки) ---
    assignee_users = set()
    for req in qs.select_related('assigned_to').prefetch_related('assignees__user'):
        if req.assigned_to:
            assignee_users.add(req.assigned_to)
        for assignee in req.assignees.all():
            assignee_users.add(assignee.user)

    worker_stats = []
    total_completed_closed = completed_closed
    for user_obj in assignee_users:
        completed_count = ServiceRequest.objects.filter(
            status__in=['completed', 'closed'],
            created_at__year=year,
            created_at__month=month
        ).filter(Q(assigned_to=user_obj) | Q(assignees__user=user_obj)).distinct().count()
        total_assigned = ServiceRequest.objects.filter(
            created_at__year=year,
            created_at__month=month
        ).filter(Q(assigned_to=user_obj) | Q(assignees__user=user_obj)).distinct().count()
        percent = (completed_count / total_completed_closed * 100) if total_completed_closed > 0 else 0
        worker_stats.append({
            'name': user_obj.get_full_name() or user_obj.username,
            'completed': completed_count,
            'total_assigned': total_assigned,
            'percent': round(percent, 1)
        })
    worker_stats.sort(key=lambda x: x['completed'], reverse=True)

    # --- Просроченные заявки по исполнителям (с ID) ---
    today = date.today()
    overdue_requests_qs = ServiceRequest.objects.filter(
        planned_date__lt=today,
        status__in=['new', 'in_progress', 'suspended']
    )
    assignee_overdue_count = {}
    for req in overdue_requests_qs.select_related('assigned_to').prefetch_related('assignees__user'):
        if req.assigned_to:
            assignee_overdue_count[req.assigned_to] = assignee_overdue_count.get(req.assigned_to, 0) + 1
        for assignee in req.assignees.all():
            user_obj = assignee.user
            assignee_overdue_count[user_obj] = assignee_overdue_count.get(user_obj, 0) + 1
    total_overdue = sum(assignee_overdue_count.values())
    overdue_assignee_stats = []
    for user_obj, cnt in assignee_overdue_count.items():
        percent = round(cnt / total_overdue * 100, 1) if total_overdue else 0
        overdue_assignee_stats.append({
            'assignee_name': user_obj.get_full_name() or user_obj.username,
            'assignee_id': user_obj.id,
            'overdue_count': cnt,
            'percent': percent,
        })
    overdue_assignee_stats.sort(key=lambda x: x['overdue_count'], reverse=True)
    

    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    available_months = [(m, month_names_ru[m]) for m in range(1, 13)]
    current_month_name = month_names_ru[month]

    context = {
        'selected_year': year,
        'selected_month': month,
        'current_month_name': current_month_name,
        'available_years': range(date.today().year - 2, date.today().year + 1),
        'available_months': available_months,
        'total_requests': total_requests,
        'completed_closed': completed_closed,
        'in_progress': in_progress,
        'overdue': overdue,
        'completion_rate': completion_rate,
        'day_labels': days_list,
        'day_created': day_created,
        'day_completed': day_completed,
        'month_labels': month_labels,
        'monthly_created': monthly_created,
        'status_labels': status_labels,
        'status_data': status_data,
        'type_labels': type_labels,
        'type_data': type_data,
        'priority_labels': priority_labels,
        'priority_data': priority_data,
        'top_executors': top_executors,
        'assignee_list': assignee_list,
        'assignee_daily': assignee_daily,
        'avg_completion_hours': round(avg_hours, 1),
        'worker_stats': worker_stats,
        'overdue_assignee_stats': overdue_assignee_stats,
    }
    return context


# ========== ОСНОВНЫЕ ПРЕДСТАВЛЕНИЯ (CRUD) ==========

@login_required
def request_list(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    qs = ServiceRequest.objects.select_related('building', 'section', 'created_by', 'assigned_to')
    if role == UserRole.WORKER:
        qs = qs.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    elif role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        qs = qs.filter(created_by=user)

    status = request.GET.get('status')
    executor = request.GET.get('executor')
    priority = request.GET.get('priority')
    search = request.GET.get('search')
    if status:
        qs = qs.filter(status=status)
    if executor:
        qs = qs.filter(assigned_to_id=executor)
    if priority:
        qs = qs.filter(priority=priority)
    if search:
        qs = qs.filter(Q(request_number__icontains=search) | Q(description__icontains=search))

    executors = User.objects.filter(profile__role=UserRole.WORKER).order_by('username')
    status_choices = ServiceRequest.STATUS_CHOICES
    priority_choices = ServiceRequest.PRIORITY_CHOICES

    paginator = Paginator(qs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    all_users = User.objects.filter(is_active=True).values(
        'id', 'username', 'first_name', 'last_name'
    ).order_by('username')

    context = {
        'requests': page_obj,
        'status_choices': status_choices,
        'executors': executors,
        'priority_choices': priority_choices,
        'selected_status': status,
        'selected_executor': executor,
        'selected_priority': priority,
        'search_query': search,
        'all_users': list(all_users),
        'user_role': role,
    }
    return render(request, 'requests_app/request_list.html', context)

@login_required
def request_create(request):
    initial = {}
    building_id = request.GET.get('building')
    if building_id and building_id.isdigit():
        try:
            building = Building.objects.get(pk=int(building_id))
            initial['building'] = building
        except Building.DoesNotExist:
            pass
    else:
        req_settings = RequestSettings.objects.first()
        if req_settings and req_settings.default_building:
            initial['building'] = req_settings.default_building

    if request.method == 'POST':
        form = ServiceRequestForm(request.user, request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req_settings = RequestSettings.objects.first()
            if req_settings and req_settings.single_building and req_settings.default_building:
                req.building = req_settings.default_building
            else:
                req.building = form.cleaned_data.get('building')
            req.created_by = request.user
            req.status = 'new'
            req.created_at = timezone.now()
            req.save()
            files = request.FILES.getlist('files')
            for f in files:
                RequestFile.objects.create(
                    request=req,
                    file=f,
                    uploaded_by=request.user,
                    description=''
                )
            messages.success(request, f'Заявка {req.request_number} успешно создана.')
            return redirect('requests_app:request_detail', pk=req.pk)
        else:
            messages.error(request, 'Ошибка в форме.')
    else:
        form = ServiceRequestForm(request.user, initial=initial)
    return render(request, 'requests_app/request_form.html', {'form': form, 'title': 'Создание заявки'})

@login_required
def request_edit(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if req.status == 'closed' and role != UserRole.ADMIN:
        messages.error(request, 'Только администратор может редактировать закрытую заявку.')
        return redirect('requests_app:request_detail', pk=pk)
    if not (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user):
        messages.error(request, 'Нет прав на редактирование.')
        return redirect('requests_app:request_detail', pk=pk)

    if request.method == 'POST':
        form = ServiceRequestForm(request.user, request.POST, instance=req)
        if form.is_valid():
            req = form.save(commit=False)
            req.save()
            delete_files = request.POST.getlist('delete_files')
            for file_id in delete_files:
                try:
                    file_obj = RequestFile.objects.get(id=file_id, request=req)
                    file_obj.file.delete()
                    file_obj.delete()
                except RequestFile.DoesNotExist:
                    pass
            new_files = request.FILES.getlist('files')
            for f in new_files:
                RequestFile.objects.create(
                    request=req,
                    file=f,
                    uploaded_by=request.user,
                    description=''
                )
            messages.success(request, 'Заявка обновлена.')
            return redirect('requests_app:request_detail', pk=req.pk)
    else:
        form = ServiceRequestForm(request.user, instance=req)
    return render(request, 'requests_app/request_form.html', {
        'form': form,
        'title': 'Редактирование заявки',
        'is_edit': True,
        'request_obj': req,
        'files': req.files.all(),
    })

@login_required
def request_detail(request, pk):
    req = get_object_or_404(ServiceRequest.objects.select_related('building', 'section', 'created_by', 'assigned_to'), pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if role == UserRole.WORKER:
            if not (req.assigned_to == user or req.assignees.filter(user=user).exists()):
                messages.error(request, 'Нет доступа.')
                return redirect('requests_app:request_list')
        else:
            if req.created_by != user:
                messages.error(request, 'Нет доступа.')
                return redirect('requests_app:request_list')
    materials_formset = UsedMaterialFormSet(instance=req)
    can_assign = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status in ['new', 'in_progress']
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    can_mark_completed = (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) and req.status == 'in_progress'
    can_suspend = (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) and req.status == 'in_progress'
    can_resume = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status == 'suspended'
    can_close = role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] and req.status == 'completed'
    can_edit = (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user) and not (req.status == 'closed' and role != UserRole.ADMIN)
    history = req.history.all()[:30]
    assignees = req.assignees.all()
    context = {
        'req': req,
        'request_obj': req,
        'materials_formset': materials_formset,
        'can_assign': can_assign,
        'can_mark_completed': can_mark_completed,
        'can_suspend': can_suspend,
        'can_resume': can_resume,
        'can_close': can_close,
        'can_edit': can_edit,
        'attachments': [],
        'history': history,
        'files': req.files.all(),
        'assignees': assignees,
    }
    return render(request, 'requests_app/request_detail.html', context)

@login_required
def request_delete(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if not (role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.created_by == user):
        messages.error(request, 'Нет прав на удаление.')
        return redirect('requests_app:request_list')
    if request.method == 'POST':
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка удалена',
        )
        req.delete()
        messages.success(request, f'Заявка {req.request_number} удалена.')
        return redirect('requests_app:request_list')
    return render(request, 'requests_app/request_confirm_delete.html', {'req': req})

@login_required
def request_assign(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER] or req.status not in ['new', 'in_progress']:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав или неверный статус'})
        messages.error(request, 'Нет прав для назначения.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')
        if assigned_to_id:
            req.assigned_to_id = assigned_to_id
            if req.status == 'new':
                req.status = 'in_progress'
            req.save()
            executor_name = req.assigned_to.get_full_name() or req.assigned_to.username
            RequestHistory.objects.create(
                request=req,
                user=request.user,
                action=f'Назначен исполнитель: {executor_name}',
            )
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Исполнитель назначен'})
            messages.success(request, 'Исполнитель назначен.')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Выберите исполнителя'})
            messages.error(request, 'Выберите исполнителя.')
        return redirect('requests_app:request_detail', pk=pk)
    return redirect('requests_app:request_detail', pk=pk)

@login_required
def request_mark_completed(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    if not (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) or req.status != 'in_progress':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав или неверный статус'})
        messages.error(request, 'Нет прав для отметки выполнения.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        req.status = 'completed'
        req.completed_date = timezone.now()
        time_spent = request.POST.get('time_spent')
        if time_spent and time_spent.isdigit():
            req.time_spent = int(time_spent)
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка отмечена как выполненная' + (f' (время: {time_spent} мин)' if time_spent else '')
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Заявка отмечена выполненной'})
        messages.success(request, 'Заявка выполнена.')
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_mark_completed.html', {'req': req})

@login_required
def request_suspend(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    is_executor = (req.assigned_to == user or req.assignees.filter(user=user).exists())
    if not (is_executor or role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]) or req.status != 'in_progress':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для приостановки'})
        messages.error(request, 'Нет прав для приостановки.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        reason = request.POST.get('suspension_reason', '').strip()
        if not reason:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Укажите причину приостановки'})
            messages.error(request, 'Укажите причину приостановки.')
        else:
            req.status = 'suspended'
            req.suspension_reason = reason
            req.save()
            RequestHistory.objects.create(
                request=req,
                user=request.user,
                action=f'Заявка приостановлена. Причина: {reason}',
            )
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Заявка приостановлена'})
            messages.success(request, f'Заявка №{req.request_number} приостановлена.')
            return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_suspend.html', {'req': req})

@login_required
def request_resume(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для возобновления'})
        messages.error(request, 'Нет прав для возобновления заявки.')
        return redirect('requests_app:request_detail', pk=pk)
    allowed_statuses = ['suspended']
    if role == UserRole.ADMIN:
        allowed_statuses.append('closed')
    if req.status not in allowed_statuses:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Заявку нельзя возобновить'})
        messages.error(request, 'Заявку нельзя возобновить.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        old_status = req.status
        req.status = 'in_progress'
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка возобновлена' + (' (после закрытия)' if old_status == 'closed' else ''),
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Заявка возобновлена'})
        messages.success(request, f'Заявка №{req.request_number} возобновлена.')
        return redirect('requests_app:request_detail', pk=pk)
    return render(request, 'requests_app/request_resume.html', {'req': req})

@login_required
def request_close(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Нет прав для закрытия'})
        messages.error(request, 'Нет прав для закрытия заявки.')
        return redirect('requests_app:request_detail', pk=pk)
    if req.status != 'completed':
        messages.error(request, 'Закрыть можно только выполненную заявку.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        material_ids = request.POST.getlist('material_id[]')
        quantities = request.POST.getlist('material_quantity[]')
        units = request.POST.getlist('material_unit[]')
        prices = request.POST.getlist('material_price[]')
        materials_used = False
        with transaction.atomic():
            for mat_id, qty_str, unit, price_str in zip(material_ids, quantities, units, prices):
                if not mat_id or not qty_str:
                    continue
                try:
                    qty = Decimal(qty_str.replace(',', '.'))
                    if qty <= 0:
                        continue
                except (ValueError, TypeError):
                    continue
                try:
                    material = Material.objects.get(pk=int(mat_id))
                except (Material.DoesNotExist, ValueError, TypeError):
                    messages.error(request, f'Материал с ID {mat_id} не найден.')
                    return redirect('requests_app:request_close', pk=req.pk)
                if material.quantity_in_stock < qty:
                    messages.error(request, f'Недостаточно материала "{material.name}" на складе (доступно: {material.quantity_in_stock} {material.unit})')
                    return redirect('requests_app:request_close', pk=req.pk)
                UsedMaterial.objects.create(
                    request=req,
                    material=material,
                    name=material.name,
                    quantity=qty,
                    unit=unit,
                    price_per_unit=Decimal(price_str) if price_str else Decimal(0)
                )
                material.quantity_in_stock -= qty
                material.save()
                materials_used = True
        req.status = 'closed'
        req.save()
        RequestHistory.objects.create(
            request=req,
            user=request.user,
            action='Заявка закрыта' + (' (с материалами)' if materials_used else ' (без материалов)')
        )
        messages.success(request, f'Заявка #{req.request_number} закрыта.')
        return redirect('requests_app:request_detail', pk=req.pk)
    materials_qs = Material.objects.all().values('id', 'name', 'unit', 'default_price')
    materials_json = list(materials_qs)
    context = {
        'request_obj': req,
        'req': req,
        'materials': materials_qs,
        'materials_json': materials_json,
    }
    return render(request, 'requests_app/request_close.html', context)


# ==================== ДАШБОРД И ОТЧЁТЫ ====================

@login_required
def request_dashboard(request):
    from users.models import UserRole
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'У вас нет доступа к дашборду.')
        return redirect('requests_app:request_list')

    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    context = get_dashboard_context(year, month)
    # Для совместимости со старым шаблоном (если требуется) можно добавить all_widgets, но они не используются
    context['all_widgets'] = []
    context['visible_widgets'] = []
    return render(request, 'requests_app/dashboard.html', context)

@login_required
def export_requests_excel(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    qs = ServiceRequest.objects.select_related('building', 'section', 'created_by', 'assigned_to')
    if role == UserRole.WORKER:
        qs = qs.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    elif role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        qs = qs.filter(created_by=user)
    status = request.GET.get('status')
    executor = request.GET.get('executor')
    priority = request.GET.get('priority')
    search = request.GET.get('search')
    if status:
        qs = qs.filter(status=status)
    if executor:
        qs = qs.filter(assigned_to_id=executor)
    if priority:
        qs = qs.filter(priority=priority)
    if search:
        qs = qs.filter(Q(request_number__icontains=search) | Q(description__icontains=search))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = ['№ заявки', 'Здание', 'Часть здания', 'Помещение', 'Тип', 'Описание', 'Приоритет', 'Статус',
               'Создатель', 'Ответственный', 'Плановая дата', 'Дата выполнения', 'Дата создания']
    ws.append(headers)
    for req in qs:
        created_by_name = ''
        if req.created_by:
            created_by_name = req.created_by.get_full_name() or req.created_by.username
        elif req.contact_name:
            created_by_name = req.contact_name
        else:
            created_by_name = 'Публичная'
        assigned_to_name = req.assigned_to.get_full_name() if req.assigned_to else ''
        ws.append([
            req.request_number,
            str(req.building),
            req.section.name if req.section else '',
            req.room_number or '',
            req.request_type.name if req.request_type else '',
            req.description[:100] if req.description else '',
            req.get_priority_display(),
            req.get_status_display(),
            created_by_name,
            assigned_to_name,
            req.planned_date.strftime('%d.%m.%Y') if req.planned_date else '',
            req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else '',
            req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="requests_export.xlsx"'
    wb.save(response)
    return response

@login_required
def custom_report(request):
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'У вас нет доступа к отчётам.')
        return redirect('requests_app:request_list')
    form = ReportForm(request.GET or None)
    if role == UserRole.WORKER:
        qs = ServiceRequest.objects.filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    else:
        qs = ServiceRequest.objects.all()
    if request.GET and form.is_valid():
        if form.cleaned_data.get('status'):
            qs = qs.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('priority'):
            qs = qs.filter(priority=form.cleaned_data['priority'])
        if form.cleaned_data.get('building'):
            qs = qs.filter(building=form.cleaned_data['building'])
        if form.cleaned_data.get('request_type'):
            qs = qs.filter(request_type=form.cleaned_data['request_type'])
        if form.cleaned_data.get('assigned_to'):
            qs = qs.filter(assigned_to=form.cleaned_data['assigned_to'])
        if form.cleaned_data.get('created_by'):
            qs = qs.filter(created_by=form.cleaned_data['created_by'])
        if form.cleaned_data.get('room_number'):
            qs = qs.filter(room_number__icontains=form.cleaned_data['room_number'])
        if form.cleaned_data.get('date_from'):
            qs = qs.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            qs = qs.filter(created_at__date__lte=form.cleaned_data['date_to'])
    status_stats = qs.values('status').annotate(count=Count('id')).order_by('status')
    status_labels = []
    status_data = []
    status_display = dict(ServiceRequest.STATUS_CHOICES)
    for item in status_stats:
        status_labels.append(status_display.get(item['status'], item['status']))
        status_data.append(item['count'])
    columns = form.cleaned_data.get('columns') if form.is_valid() else []
    if not columns:
        columns = ['request_number', 'building', 'priority', 'status', 'created_by', 'assigned_to', 'created_at']
    field_map = {
        'request_number': '№ заявки',
        'building': 'Здание',
        'section': 'Часть здания',
        'room_number': 'Помещение',
        'request_type': 'Тип',
        'description': 'Описание',
        'priority': 'Приоритет',
        'status': 'Статус',
        'created_by': 'Создатель',
        'assigned_to': 'Ответственный',
        'planned_date': 'Плановая дата',
        'completed_date': 'Дата выполнения',
        'created_at': 'Дата создания',
        'comment': 'Комментарий',
    }
    if request.GET.get('export') == '1':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт по заявкам"
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_map.get(col, col))
            cell.font = openpyxl.styles.Font(bold=True)
        for row_idx, req in enumerate(qs.select_related('building', 'section', 'created_by', 'assigned_to', 'request_type'), 2):
            for col_idx, col in enumerate(columns, 1):
                value = ''
                if col == 'request_number':
                    value = req.request_number
                elif col == 'building':
                    value = str(req.building)
                elif col == 'section':
                    value = req.section.name if req.section else ''
                elif col == 'room_number':
                    value = req.room_number or ''
                elif col == 'request_type':
                    value = req.request_type.name if req.request_type else ''
                elif col == 'description':
                    value = req.description[:200] if req.description else ''
                elif col == 'priority':
                    value = req.get_priority_display()
                elif col == 'status':
                    value = req.get_status_display()
                elif col == 'created_by':
                    value = req.created_by.get_full_name() if req.created_by else (req.contact_name or 'Публичная')
                elif col == 'assigned_to':
                    value = req.assigned_to.get_full_name() if req.assigned_to else ''
                elif col == 'planned_date':
                    value = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
                elif col == 'completed_date':
                    value = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
                elif col == 'created_at':
                    value = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
                elif col == 'comment':
                    value = req.comment[:200] if req.comment else ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        for col_idx in range(1, len(columns) + 1):
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="custom_report.xlsx"'
        wb.save(response)
        return response
    data = []
    for req in qs.select_related('building', 'section', 'created_by', 'assigned_to', 'request_type'):
        row = {}
        for col in columns:
            if col == 'request_number':
                row[col] = req.request_number
            elif col == 'building':
                row[col] = str(req.building)
            elif col == 'section':
                row[col] = req.section.name if req.section else ''
            elif col == 'room_number':
                row[col] = req.room_number or ''
            elif col == 'request_type':
                row[col] = req.request_type.name if req.request_type else ''
            elif col == 'description':
                row[col] = req.description[:100] if req.description else ''
            elif col == 'priority':
                row[col] = req.get_priority_display()
            elif col == 'status':
                row[col] = req.get_status_display()
            elif col == 'created_by':
                row[col] = req.created_by.get_full_name() if req.created_by else (req.contact_name or 'Публичная')
            elif col == 'assigned_to':
                row[col] = req.assigned_to.get_full_name() if req.assigned_to else ''
            elif col == 'planned_date':
                row[col] = req.planned_date.strftime('%d.%m.%Y') if req.planned_date else ''
            elif col == 'completed_date':
                row[col] = req.completed_date.strftime('%d.%m.%Y %H:%M') if req.completed_date else ''
            elif col == 'created_at':
                row[col] = req.created_at.strftime('%d.%m.%Y %H:%M') if req.created_at else ''
            elif col == 'comment':
                row[col] = req.comment[:200] if req.comment else ''
            else:
                row[col] = ''
        data.append(row)
    context = {
        'form': form,
        'data': data,
        'columns': columns,
        'column_labels': field_map,
        'field_map': field_map,
        'status_labels': status_labels,
        'status_data': status_data,
    }
    return render(request, 'requests_app/custom_report.html', context)


# ==================== УПРАВЛЕНИЕ МАТЕРИАЛАМИ ====================

@login_required
def material_stock(request):
    search_query = request.GET.get('search', '').strip()
    materials_qs = Material.objects.all().order_by('name')
    if search_query:
        materials_qs = materials_qs.filter(name__icontains=search_query)
    paginator = Paginator(materials_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'requests_app/material_stock.html', {
        'materials': page_obj,
        'search': search_query,
    })

@login_required
def material_add(request):
    role = request.user.profile.role if hasattr(request.user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для добавления материалов.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            messages.success(request, f'Материал "{material.name}" добавлен.')
            return redirect('requests_app:material_stock')
    else:
        form = MaterialForm()
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Добавить материал'})

@login_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для редактирования материала.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, f'Материал "{material.name}" успешно обновлён.')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме.')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'requests_app/material_form.html', {'form': form, 'title': 'Редактировать материал'})

@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для удаления материала.')
        return redirect('requests_app:material_stock')
    if request.method == 'POST':
        name = material.name
        material.delete()
        messages.success(request, f'Материал "{name}" удалён.')
        return redirect('requests_app:material_stock')
    return render(request, 'requests_app/material_confirm_delete.html', {'material': material})

@login_required
def material_delete_ajax(request, pk):
    material = get_object_or_404(Material, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        return JsonResponse({'success': False, 'error': 'Нет прав'}, status=403)
    if request.method == 'POST':
        material.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)

@login_required
def material_stock_export(request):
    materials = Material.objects.all().values('name', 'unit', 'quantity_in_stock', 'default_price')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Склад материалов"
    headers = ['Наименование', 'Единица измерения', 'Количество на складе', 'Цена за единицу']
    ws.append(headers)
    for m in materials:
        ws.append([
            m['name'],
            m['unit'],
            float(m['quantity_in_stock']),
            float(m['default_price'])
        ])
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="material_stock_export.xlsx"'
    wb.save(response)
    return response

@login_required
def import_materials(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        form = ImportMaterialsForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            created = 0
            updated = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0] if row[0] else None
                unit = row[1] if len(row) > 1 else None
                default_price = row[2] if len(row) > 2 and row[2] is not None else 0
                quantity_in_stock = row[3] if len(row) > 3 and row[3] is not None else 0
                if name and unit:
                    try:
                        default_price = float(str(default_price).replace(',', '.'))
                    except (ValueError, TypeError):
                        default_price = 0.0
                    try:
                        quantity_in_stock = float(str(quantity_in_stock).replace(',', '.'))
                    except (ValueError, TypeError):
                        quantity_in_stock = 0.0
                    material, is_created = Material.objects.update_or_create(
                        name=name,
                        defaults={
                            'unit': unit,
                            'default_price': default_price,
                            'quantity_in_stock': quantity_in_stock
                        }
                    )
                    if is_created:
                        created += 1
                    else:
                        updated += 1
            messages.success(request, f'Импортировано: добавлено {created}, обновлено {updated}.')
            return redirect('requests_app:material_stock')
        else:
            messages.error(request, 'Ошибка в форме. Проверьте файл.')
    else:
        form = ImportMaterialsForm()
    return render(request, 'requests_app/import_materials.html', {'form': form})

@login_required
def download_materials_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Материалы"
    headers = ['name', 'unit', 'default_price', 'quantity_in_stock']
    ws.append(headers)
    ws.append(['Краска', 'л', 350.00, 100])
    ws.append(['Лампа светодиодная', 'шт', 450.00, 50])
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="materials_import_template.xlsx"'
    wb.save(response)
    return response


# ==================== УПРАВЛЕНИЕ ИСПОЛНИТЕЛЯМИ ====================

@login_required
def request_add_assignee(request, pk):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для назначения исполнителей.')
        return redirect('requests_app:request_detail', pk=pk)
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        if user_id:
            try:
                assignee = User.objects.get(pk=user_id)
                obj, created = RequestAssignee.objects.get_or_create(request=req, user=assignee)
                if created:
                    RequestHistory.objects.create(
                        request=req,
                        user=request.user,
                        action=f'Добавлен исполнитель: {assignee.get_full_name() or assignee.username}'
                    )
                    messages.success(request, f'Исполнитель {assignee.get_full_name() or assignee.username} добавлен.')
                else:
                    messages.warning(request, 'Этот исполнитель уже назначен.')
            except User.DoesNotExist:
                messages.error(request, 'Пользователь не найден.')
        else:
            messages.error(request, 'Выберите пользователя.')
        return redirect('requests_app:request_detail', pk=pk)
    assigned_user_ids = req.assignees.values_list('user_id', flat=True)
    available_users = User.objects.filter(is_active=True).exclude(id__in=assigned_user_ids).exclude(id=req.assigned_to_id).order_by('username')
    return render(request, 'requests_app/add_assignee.html', {'request_obj': req, 'users': available_users})

@login_required
def request_remove_assignee(request, pk, user_id):
    req = get_object_or_404(ServiceRequest, pk=pk)
    user = request.user
    role = user.profile.role if hasattr(user, 'profile') else None
    if role not in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]:
        messages.error(request, 'Нет прав для удаления исполнителей.')
        return redirect('requests_app:request_detail', pk=pk)
    assignee = get_object_or_404(RequestAssignee, request=req, user_id=user_id)
    assignee_name = assignee.user.get_full_name() or assignee.user.username
    assignee.delete()
    RequestHistory.objects.create(
        request=req,
        user=request.user,
        action=f'Удалён исполнитель: {assignee_name}'
    )
    messages.success(request, f'Исполнитель {assignee_name} удалён.')
    return redirect('requests_app:request_detail', pk=pk)


# ==================== ПУБЛИЧНАЯ ФОРМА ====================

@csrf_protect
@never_cache
@rate_limit()
def public_request_create(request):
    lang = request.GET.get('lang', 'ru')
    if lang not in ['ru', 'en']:
        lang = 'ru'

    req_settings = RequestSettings.objects.first()
    single_building_mode = req_settings and req_settings.single_building and req_settings.default_building
    default_building = req_settings.default_building if single_building_mode else None

    section_queryset = None
    sections_list = []
    if single_building_mode and default_building:
        if hasattr(default_building, 'sections'):
            section_queryset = default_building.sections.all().order_by('name')
            sections_list = list(section_queryset)
        else:
            section_queryset = default_building.buildingsection_set.all().order_by('name')
            sections_list = list(section_queryset)

    if request.method == 'POST':
        form = PublicRequestForm(request.POST, request.FILES, lang=lang, section_queryset=section_queryset)

        if form.is_valid():
            files = request.FILES.getlist('files')
            max_files = 5
            max_size = 5 * 1024 * 1024
            allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'application/pdf']
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.pdf']

            if len(files) > max_files:
                messages.error(request, f'Можно прикрепить не более {max_files} файлов.')
                post_data = request.POST.copy()
                post_data.update(generate_new_captcha(lang))
                form = PublicRequestForm(post_data, request.FILES, lang=lang, section_queryset=section_queryset)
                context = {
                    'form': form,
                    'lang': lang,
                    'hide_navbar': True,
                    'single_building_mode': single_building_mode,
                    'sections_list': sections_list,
                }
                return render(request, f'requests_app/public_request_form_{lang}.html', context)

            for f in files:
                if f.size > max_size:
                    messages.error(request, f'Файл "{f.name}" превышает 5 МБ.')
                    post_data = request.POST.copy()
                    post_data.update(generate_new_captcha(lang))
                    form = PublicRequestForm(post_data, request.FILES, lang=lang, section_queryset=section_queryset)
                    context = {
                        'form': form,
                        'lang': lang,
                        'hide_navbar': True,
                        'single_building_mode': single_building_mode,
                        'sections_list': sections_list,
                    }
                    return render(request, f'requests_app/public_request_form_{lang}.html', context)

                ext = os.path.splitext(f.name)[1].lower()
                if ext not in allowed_extensions:
                    messages.error(request, f'Файл "{f.name}" имеет недопустимое расширение. Разрешены: {", ".join(allowed_extensions)}')
                    post_data = request.POST.copy()
                    post_data.update(generate_new_captcha(lang))
                    form = PublicRequestForm(post_data, request.FILES, lang=lang, section_queryset=section_queryset)
                    context = {
                        'form': form,
                        'lang': lang,
                        'hide_navbar': True,
                        'single_building_mode': single_building_mode,
                        'sections_list': sections_list,
                    }
                    return render(request, f'requests_app/public_request_form_{lang}.html', context)

                try:
                    import magic
                    f.seek(0)
                    mime = magic.from_buffer(f.read(1024), mime=True)
                    f.seek(0)
                    if mime not in allowed_types:
                        messages.error(request, f'Файл "{f.name}" имеет недопустимый тип ({mime}).')
                        post_data = request.POST.copy()
                        post_data.update(generate_new_captcha(lang))
                        form = PublicRequestForm(post_data, request.FILES, lang=lang, section_queryset=section_queryset)
                        context = {
                            'form': form,
                            'lang': lang,
                            'hide_navbar': True,
                            'single_building_mode': single_building_mode,
                            'sections_list': sections_list,
                        }
                        return render(request, f'requests_app/public_request_form_{lang}.html', context)
                except ImportError:
                    pass

            sr = form.save(commit=False)
            sr.created_by = None
            sr.assigned_to = None
            sr.status = 'new'
            sr.priority = 'low'
            sr.ip_address = request.META.get('REMOTE_ADDR')

            if single_building_mode:
                sr.building = default_building
                section_id = request.POST.get('section')
                if section_id:
                    try:
                        sr.section = BuildingSection.objects.get(id=section_id)
                    except BuildingSection.DoesNotExist:
                        pass
            else:
                sr.building = form.cleaned_data.get('building')
                sr.section = form.cleaned_data.get('section')

            if lang == 'en' and sr.description:
                sr.description = translate_to_russian(sr.description)

            contact_info = f"Контактное лицо: {sr.contact_name or 'не указано'}, Телефон: {sr.contact_phone or 'не указан'}"
            if sr.comment:
                sr.comment = f"{sr.comment}\n{contact_info}"
            else:
                sr.comment = contact_info

            sr.save()

            msg = (
                f"🔔 <b>Новая публичная заявка!</b>\n"
                f"<b>Номер:</b> {sr.request_number}\n"
                f"<b>Здание:</b> {sr.building}\n"
                f"<b>Помещение:</b> {sr.room_number or '—'}\n"
                f"<b>Тип:</b> {sr.request_type.name}\n"
                f"<b>Описание:</b> {sr.description[:100]}\n"
            )
            if sr.contact_name or sr.contact_phone:
                msg += f"<b>Контакты:</b> {sr.contact_name or ''} {sr.contact_phone or ''}\n"
            msg += f"<a href='https://exploitationonline.ru/requests/{sr.id}/'>Перейти к заявке</a>"
            send_telegram_notification(msg)

            for f in files:
                RequestFile.objects.create(
                    request=sr,
                    file=f,
                    uploaded_by=None,
                    description='Загружено из публичной формы'
                )

            if settings.DEFAULT_FROM_EMAIL and hasattr(settings, 'ADMINS') and settings.ADMINS:
                try:
                    subject = f"New public request #{sr.request_number}"
                    html_message = render_to_string('requests_app/email_new_public_request.html', {'request': sr})
                    send_mail(
                        subject,
                        f"Request #{sr.request_number} from {sr.contact_name or 'Anonymous'}",
                        settings.DEFAULT_FROM_EMAIL,
                        [email for name, email in settings.ADMINS],
                        fail_silently=True,
                        html_message=html_message
                    )
                except Exception:
                    pass

            now = timezone.localtime(timezone.now())
            off_hours = False
            if now.weekday() >= 5:
                off_hours = True
            else:
                work_start = time(9, 30)
                work_end = time(18, 0)
                current_time = now.time()
                if current_time < work_start or current_time >= work_end:
                    off_hours = True

            off_param = '&off_hours=1' if off_hours else ''
            return redirect(f'{reverse("requests_app:public_request_success")}?lang={lang}{off_param}')

        else:
            post_data = request.POST.copy()
            post_data.update(generate_new_captcha(lang))
            form = PublicRequestForm(post_data, request.FILES, lang=lang, section_queryset=section_queryset)
            context = {
                'form': form,
                'lang': lang,
                'hide_navbar': True,
                'single_building_mode': single_building_mode,
                'sections_list': sections_list,
            }
            return render(request, f'requests_app/public_request_form_{lang}.html', context)

    else:
        form = PublicRequestForm(lang=lang, section_queryset=section_queryset)
        context = {
            'form': form,
            'lang': lang,
            'hide_navbar': True,
            'single_building_mode': single_building_mode,
            'sections_list': sections_list,
        }
        return render(request, f'requests_app/public_request_form_{lang}.html', context)

def public_request_success(request):
    lang = request.GET.get('lang', 'ru')
    if lang not in ['ru', 'en']:
        lang = 'ru'
    off_hours = request.GET.get('off_hours') == '1'
    return render(request, f'requests_app/public_request_success_{lang}.html', {'hide_navbar': True, 'off_hours': off_hours})


# ==================== API ДЛЯ ДИНАМИЧЕСКОЙ ЗАГРУЗКИ СЕКЦИЙ ====================

@login_required
def api_building_sections(request):
    building_id = request.GET.get('building_id')
    if not building_id:
        return JsonResponse([], safe=False)
    try:
        building = Building.objects.get(pk=building_id)
        sections = building.sections.all().values('id', 'name')
        return JsonResponse(list(sections), safe=False)
    except Building.DoesNotExist:
        return JsonResponse([], safe=False)
    
@login_required
def api_overdue_requests(request):
    """Возвращает JSON со списком просроченных заявок для указанного исполнителя."""
    from datetime import date
    from django.db.models import Q
    
    assignee_id = request.GET.get('assignee_id')
    if not assignee_id or not assignee_id.isdigit():
        return JsonResponse({'error': 'Не указан ID исполнителя'}, status=400)
    
    try:
        user = User.objects.get(pk=int(assignee_id))
    except User.DoesNotExist:
        return JsonResponse({'error': 'Исполнитель не найден'}, status=404)
    
    today = date.today()
    overdue_requests = ServiceRequest.objects.filter(
        planned_date__lt=today,
        status__in=['new', 'in_progress', 'suspended']
    ).filter(Q(assigned_to=user) | Q(assignees__user=user)).distinct()
    
    requests_data = []
    for req in overdue_requests:
        requests_data.append({
            'id': req.id,
            'request_number': req.request_number,
            'description': req.description,
            'planned_date': req.planned_date.strftime('%d.%m.%Y') if req.planned_date else '',
            'status_display': req.get_status_display(),
        })
    
    return JsonResponse({'requests': requests_data})


# ==================== ПОЛНЫЙ БЭКАП ЗАЯВОК (Excel) ====================

@login_required
def export_requests_full_backup(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role in ['ADMIN', 'ENGINEER', 'DISPATCHER'])):
        messages.error(request, 'У вас нет прав на экспорт бэкапа заявок.')
        return redirect('requests_app:request_list')

    wb = openpyxl.Workbook()
    ws_requests = wb.active
    ws_requests.title = "Заявки"
    request_headers = [
        'request_number', 'building_address', 'section_name', 'room_number', 'request_type_name',
        'description', 'priority', 'status', 'created_by_username', 'assigned_to_username',
        'planned_date', 'completed_date', 'comment', 'created_at', 'updated_at',
        'track_time', 'time_spent', 'suspension_reason', 'contact_name', 'contact_phone', 'ip_address'
    ]
    ws_requests.append(request_headers)
    qs = ServiceRequest.objects.select_related('building', 'section', 'request_type', 'created_by', 'assigned_to').all()
    for req in qs:
        ws_requests.append([
            req.request_number,
            req.building.address if req.building else '',
            req.section.name if req.section else '',
            req.room_number or '',
            req.request_type.name if req.request_type else '',
            req.description or '',
            req.priority,
            req.status,
            req.created_by.username if req.created_by else (req.contact_name or ''),
            req.assigned_to.username if req.assigned_to else '',
            req.planned_date.strftime('%Y-%m-%d') if req.planned_date else '',
            req.completed_date.strftime('%Y-%m-%d %H:%M:%S') if req.completed_date else '',
            req.comment or '',
            req.created_at.strftime('%Y-%m-%d %H:%M:%S') if req.created_at else '',
            req.updated_at.strftime('%Y-%m-%d %H:%M:%S') if req.updated_at else '',
            req.track_time,
            req.time_spent or '',
            req.suspension_reason or '',
            req.contact_name or '',
            req.contact_phone or '',
            req.ip_address or '',
        ])

    ws_materials = wb.create_sheet("Материалы")
    material_headers = ['request_number', 'material_name', 'quantity', 'unit', 'price_per_unit', 'total_price']
    ws_materials.append(material_headers)
    used_materials = UsedMaterial.objects.select_related('request').all()
    for um in used_materials:
        ws_materials.append([
            um.request.request_number,
            um.name,
            float(um.quantity),
            um.unit,
            float(um.price_per_unit),
            float(um.total_price) if um.total_price else '',
        ])

    ws_files = wb.create_sheet("Файлы")
    file_headers = ['request_number', 'file_name', 'description', 'uploaded_by_username', 'uploaded_at']
    ws_files.append(file_headers)
    files = RequestFile.objects.select_related('request', 'uploaded_by').all()
    for f in files:
        ws_files.append([
            f.request.request_number,
            f.get_file_name(),
            f.description or '',
            f.uploaded_by.username if f.uploaded_by else '',
            f.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if f.uploaded_at else '',
        ])

    ws_assignees = wb.create_sheet("Исполнители")
    assignee_headers = ['request_number', 'user_username']
    ws_assignees.append(assignee_headers)
    assignees = RequestAssignee.objects.select_related('request', 'user').all()
    for a in assignees:
        ws_assignees.append([
            a.request.request_number,
            a.user.username,
        ])

    ws_history = wb.create_sheet("История")
    history_headers = ['request_number', 'user_username', 'action', 'old_value', 'new_value', 'created_at']
    ws_history.append(history_headers)
    history = RequestHistory.objects.select_related('request', 'user').all()
    for h in history:
        ws_history.append([
            h.request.request_number,
            h.user.username if h.user else '',
            h.action,
            h.old_value or '',
            h.new_value or '',
            h.created_at.strftime('%Y-%m-%d %H:%M:%S') if h.created_at else '',
        ])

    for ws in [ws_requests, ws_materials, ws_files, ws_assignees, ws_history]:
        for col in ws.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    filename = f"requests_full_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def import_requests_full_backup(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role in ['ADMIN', 'ENGINEER', 'DISPATCHER'])):
        messages.error(request, 'У вас нет прав на импорт бэкапа заявок.')
        return redirect('requests_app:request_list')

    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        if not backup_file.name.endswith('.xlsx'):
            messages.error(request, 'Поддерживаются только файлы .xlsx')
            return redirect('requests_app:request_list')
        try:
            wb = openpyxl.load_workbook(backup_file, data_only=True)
            required_sheets = ['Заявки', 'Материалы', 'Файлы', 'Исполнители', 'История']
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    messages.error(request, f'В файле отсутствует лист "{sheet}"')
                    return redirect('requests_app:request_list')

            ws_requests = wb['Заявки']
            request_headers = [cell.value for cell in ws_requests[1]]
            req_col_idx = {h: i for i, h in enumerate(request_headers) if h}
            required_req_fields = ['request_number', 'building_address', 'request_type_name']
            for f in required_req_fields:
                if f not in req_col_idx:
                    messages.error(request, f'В листе "Заявки" отсутствует колонка "{f}"')
                    return redirect('requests_app:request_list')

            created_count = 0
            updated_count = 0
            errors = []
            request_numbers_in_backup = []
            for row in ws_requests.iter_rows(min_row=2, values_only=True):
                if row and row[req_col_idx['request_number']]:
                    request_numbers_in_backup.append(str(row[req_col_idx['request_number']]).strip())

            for row_idx, row in enumerate(ws_requests.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(c is None for c in row):
                    continue
                contract_number = str(row[req_col_idx['request_number']]).strip() if row[req_col_idx['request_number']] else None
                if not contract_number:
                    errors.append(f'Строка {row_idx}: отсутствует номер заявки')
                    continue

                building_address = str(row[req_col_idx['building_address']]).strip() if row[req_col_idx['building_address']] else None
                building = None
                if building_address:
                    building = Building.objects.filter(address__iexact=building_address).first()
                    if not building:
                        errors.append(f'Строка {row_idx}: здание "{building_address}" не найдено')
                        continue

                section_name = str(row[req_col_idx.get('section_name', 2)]).strip() if req_col_idx.get('section_name') and row[req_col_idx.get('section_name')] else None
                section = None
                if section_name and building:
                    section = BuildingSection.objects.filter(building=building, name__iexact=section_name).first()
                    if not section:
                        errors.append(f'Строка {row_idx}: часть здания "{section_name}" не найдена для здания {building.address}')

                request_type_name = str(row[req_col_idx['request_type_name']]).strip() if row[req_col_idx['request_type_name']] else None
                request_type = None
                if request_type_name:
                    request_type = RequestType.objects.filter(name__iexact=request_type_name).first()
                    if not request_type:
                        errors.append(f'Строка {row_idx}: тип заявки "{request_type_name}" не найден')
                        continue

                created_by_username = str(row[req_col_idx.get('created_by_username', 8)]).strip() if req_col_idx.get('created_by_username') and row[req_col_idx.get('created_by_username')] else None
                created_by = None
                if created_by_username:
                    created_by = User.objects.filter(username=created_by_username).first()

                assigned_to_username = str(row[req_col_idx.get('assigned_to_username', 9)]).strip() if req_col_idx.get('assigned_to_username') and row[req_col_idx.get('assigned_to_username')] else None
                assigned_to = None
                if assigned_to_username:
                    assigned_to = User.objects.filter(username=assigned_to_username).first()

                planned_date = parse_date(row[req_col_idx.get('planned_date', 10)]) if req_col_idx.get('planned_date') and row[req_col_idx.get('planned_date')] else None
                completed_date = parse_datetime(row[req_col_idx.get('completed_date', 11)]) if req_col_idx.get('completed_date') and row[req_col_idx.get('completed_date')] else None
                created_at = parse_datetime(row[req_col_idx.get('created_at', 13)]) if req_col_idx.get('created_at') and row[req_col_idx.get('created_at')] else None
                updated_at = parse_datetime(row[req_col_idx.get('updated_at', 14)]) if req_col_idx.get('updated_at') and row[req_col_idx.get('updated_at')] else None

                defaults = {
                    'building': building,
                    'section': section,
                    'room_number': str(row[req_col_idx.get('room_number', 3)]).strip() if req_col_idx.get('room_number') and row[req_col_idx.get('room_number')] else '',
                    'request_type': request_type,
                    'description': str(row[req_col_idx.get('description', 5)]).strip() if req_col_idx.get('description') and row[req_col_idx.get('description')] else '',
                    'priority': str(row[req_col_idx.get('priority', 6)]).strip() if req_col_idx.get('priority') and row[req_col_idx.get('priority')] else 'medium',
                    'status': str(row[req_col_idx.get('status', 7)]).strip() if req_col_idx.get('status') and row[req_col_idx.get('status')] else 'new',
                    'created_by': created_by,
                    'assigned_to': assigned_to,
                    'planned_date': planned_date,
                    'completed_date': completed_date,
                    'comment': str(row[req_col_idx.get('comment', 12)]).strip() if req_col_idx.get('comment') and row[req_col_idx.get('comment')] else '',
                    'created_at': created_at,
                    'updated_at': updated_at,
                    'track_time': bool(row[req_col_idx.get('track_time', 15)]) if req_col_idx.get('track_time') and row[req_col_idx.get('track_time')] else False,
                    'time_spent': int(row[req_col_idx.get('time_spent', 16)]) if req_col_idx.get('time_spent') and row[req_col_idx.get('time_spent')] else None,
                    'suspension_reason': str(row[req_col_idx.get('suspension_reason', 17)]).strip() if req_col_idx.get('suspension_reason') and row[req_col_idx.get('suspension_reason')] else '',
                    'contact_name': str(row[req_col_idx.get('contact_name', 18)]).strip() if req_col_idx.get('contact_name') and row[req_col_idx.get('contact_name')] else '',
                    'contact_phone': str(row[req_col_idx.get('contact_phone', 19)]).strip() if req_col_idx.get('contact_phone') and row[req_col_idx.get('contact_phone')] else '',
                    'ip_address': str(row[req_col_idx.get('ip_address', 20)]).strip() if req_col_idx.get('ip_address') and row[req_col_idx.get('ip_address')] else '',
                }
                try:
                    req_obj, created = ServiceRequest.objects.update_or_create(
                        request_number=contract_number,
                        defaults=defaults
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append(f'Строка {row_idx} (заявка {contract_number}): {str(e)}')

            # Импорт материалов
            if 'Материалы' in wb.sheetnames:
                ws_materials = wb['Материалы']
                material_headers = [cell.value for cell in ws_materials[1]]
                m_col_idx = {h: i for i, h in enumerate(material_headers) if h}
                if 'request_number' in m_col_idx and 'material_name' in m_col_idx:
                    UsedMaterial.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_materials.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[m_col_idx['request_number']]).strip() if row[m_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Материалы строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Материалы строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        material_name = str(row[m_col_idx['material_name']]).strip() if row[m_col_idx['material_name']] else None
                        if not material_name:
                            errors.append(f'Материалы строка {row_idx}: нет названия материала')
                            continue
                        quantity = Decimal(str(row[m_col_idx['quantity']]).replace(',', '.')) if row[m_col_idx['quantity']] else Decimal('0')
                        unit = str(row[m_col_idx['unit']]).strip() if row[m_col_idx['unit']] else ''
                        price_per_unit = Decimal(str(row[m_col_idx['price_per_unit']]).replace(',', '.')) if row[m_col_idx['price_per_unit']] else Decimal('0')
                        total_price = Decimal(str(row[m_col_idx['total_price']]).replace(',', '.')) if m_col_idx.get('total_price') and row[m_col_idx['total_price']] else quantity * price_per_unit
                        UsedMaterial.objects.create(
                            request=req_obj,
                            material=None,
                            name=material_name,
                            quantity=quantity,
                            unit=unit,
                            price_per_unit=price_per_unit,
                            total_price=total_price
                        )
                else:
                    errors.append('Лист "Материалы" не имеет обязательных колонок request_number, material_name')

            # Импорт файлов (сохраняем только информацию, сам файл не восстанавливаем)
            if 'Файлы' in wb.sheetnames:
                ws_files = wb['Файлы']
                file_headers = [cell.value for cell in ws_files[1]]
                f_col_idx = {h: i for i, h in enumerate(file_headers) if h}
                if 'request_number' in f_col_idx and 'file_name' in f_col_idx:
                    RequestFile.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_files.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[f_col_idx['request_number']]).strip() if row[f_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Файлы строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Файлы строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        file_name = str(row[f_col_idx['file_name']]).strip()
                        description = str(row[f_col_idx.get('description', 2)]).strip() if f_col_idx.get('description') and row[f_col_idx.get('description')] else ''
                        uploaded_by_username = str(row[f_col_idx.get('uploaded_by_username', 3)]).strip() if f_col_idx.get('uploaded_by_username') and row[f_col_idx.get('uploaded_by_username')] else None
                        uploaded_by = None
                        if uploaded_by_username:
                            uploaded_by = User.objects.filter(username=uploaded_by_username).first()
                        uploaded_at = parse_datetime(row[f_col_idx.get('uploaded_at', 4)]) if f_col_idx.get('uploaded_at') and row[f_col_idx.get('uploaded_at')] else None
                        RequestFile.objects.create(
                            request=req_obj,
                            file=None,
                            uploaded_by=uploaded_by,
                            uploaded_at=uploaded_at or timezone.now(),
                            description=description
                        )
                else:
                    errors.append('Лист "Файлы" не имеет обязательных колонок request_number, file_name')

            # Импорт исполнителей
            if 'Исполнители' in wb.sheetnames:
                ws_assignees = wb['Исполнители']
                assignee_headers = [cell.value for cell in ws_assignees[1]]
                a_col_idx = {h: i for i, h in enumerate(assignee_headers) if h}
                if 'request_number' in a_col_idx and 'user_username' in a_col_idx:
                    RequestAssignee.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_assignees.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[a_col_idx['request_number']]).strip() if row[a_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'Исполнители строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'Исполнители строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        username = str(row[a_col_idx['user_username']]).strip() if row[a_col_idx['user_username']] else None
                        if not username:
                            errors.append(f'Исполнители строка {row_idx}: нет имени пользователя')
                            continue
                        user_obj = User.objects.filter(username=username).first()
                        if not user_obj:
                            errors.append(f'Исполнители строка {row_idx}: пользователь {username} не найден')
                            continue
                        RequestAssignee.objects.get_or_create(request=req_obj, user=user_obj)
                else:
                    errors.append('Лист "Исполнители" не имеет обязательных колонок request_number, user_username')

            # Импорт истории
            if 'История' in wb.sheetnames:
                ws_history = wb['История']
                history_headers = [cell.value for cell in ws_history[1]]
                h_col_idx = {h: i for i, h in enumerate(history_headers) if h}
                if 'request_number' in h_col_idx and 'action' in h_col_idx:
                    RequestHistory.objects.filter(request__request_number__in=request_numbers_in_backup).delete()
                    for row_idx, row in enumerate(ws_history.iter_rows(min_row=2, values_only=True), start=2):
                        if not row or all(c is None for c in row):
                            continue
                        request_number = str(row[h_col_idx['request_number']]).strip() if row[h_col_idx['request_number']] else None
                        if not request_number:
                            errors.append(f'История строка {row_idx}: нет номера заявки')
                            continue
                        req_obj = ServiceRequest.objects.filter(request_number=request_number).first()
                        if not req_obj:
                            errors.append(f'История строка {row_idx}: заявка {request_number} не найдена')
                            continue
                        action = str(row[h_col_idx['action']]).strip() if row[h_col_idx['action']] else ''
                        if not action:
                            errors.append(f'История строка {row_idx}: нет действия')
                            continue
                        user_username = str(row[h_col_idx.get('user_username', 1)]).strip() if h_col_idx.get('user_username') and row[h_col_idx.get('user_username')] else None
                        user_obj = None
                        if user_username:
                            user_obj = User.objects.filter(username=user_username).first()
                        old_value = str(row[h_col_idx.get('old_value', 2)]).strip() if h_col_idx.get('old_value') and row[h_col_idx.get('old_value')] else ''
                        new_value = str(row[h_col_idx.get('new_value', 3)]).strip() if h_col_idx.get('new_value') and row[h_col_idx.get('new_value')] else ''
                        created_at = parse_datetime(row[h_col_idx.get('created_at', 4)]) if h_col_idx.get('created_at') and row[h_col_idx.get('created_at')] else None
                        RequestHistory.objects.create(
                            request=req_obj,
                            user=user_obj,
                            action=action,
                            old_value=old_value,
                            new_value=new_value,
                            created_at=created_at or timezone.now()
                        )
                else:
                    errors.append('Лист "История" не имеет обязательных колонок request_number, action')

            msg = f'Импорт завершён. Заявки: создано {created_count}, обновлено {updated_count}.'
            if errors:
                msg += f' Ошибки ({len(errors)}): {", ".join(errors[:10])}'
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Ошибка обработки файла: {str(e)}')
        return redirect('requests_app:request_list')
    return render(request, 'requests_app/import_requests_backup.html')