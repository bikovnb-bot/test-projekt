# energy/views.py
# Полный файл с исправлениями критических ошибок, поддержкой is_technical и историей начальных показаний

import json
from decimal import Decimal
from datetime import datetime, timedelta, date
from calendar import monthrange
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from dateutil import parser as date_parser

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, UpdateView, DeleteView
from django.http import JsonResponse, HttpResponse, Http404
from django.template.loader import render_to_string
from django.db.models import Sum
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_POST
from django.core.exceptions import PermissionDenied
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction

from .models import (
    Meter, Reading, ZoneReading, TariffComponent, ResourceType, MeterDocument,
    InitialZoneReading, InitialValueHistory, InitialZoneValueHistory
)
from .forms import ReadingForm, MeterForm, ReadingEditForm, MeterDocumentForm, ImportReadingsForm
from .utils import (
    can_view_all_meters, can_edit_all_meters, can_assign_owner,
    can_view_meter, can_edit_meter, can_delete_meter,
    can_edit_reading, can_delete_reading,
    get_avg_consumption, is_anomaly,
    log_action,
)
from users.decorators import has_contract_access, has_contract_edit_access
from users.models import UserRole
from .forms import ResetInitialReadingsForm


# ------------------------------------------------------------
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ ПРОВЕРКИ ПРАВ В ENERGY
# ------------------------------------------------------------
def _check_energy_view_access(user):
    return has_contract_access(user)

def _check_energy_edit_access(user):
    return has_contract_edit_access(user)


# ==================== ФУНКЦИЯ ДЛЯ ПОЛУЧЕНИЯ КОНТЕКСТА ДАШБОРДА ====================
def get_dashboard_context(year=None, month=None, resource_type_id='all'):
    from django.db.models import Sum
    from calendar import monthrange
    from collections import defaultdict
    from decimal import Decimal

    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    start_date = date(year, month, 1)
    end_date = start_date.replace(day=monthrange(year, month)[1])
    prev_start_date = date(year-1, month, 1)
    prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])

    # Базовый queryset активных счётчиков (исключаем технические)
    meters = Meter.objects.filter(is_active=True, is_technical=False)
    if resource_type_id != 'all':
        meters = meters.filter(resource_type_id=resource_type_id)

    total_current = Decimal('0')
    meter_count = meters.count()
    anomaly_count = 0
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            current = sum(r.total_consumption() for r in readings)
        else:
            current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        total_current += current
        avg = get_avg_consumption(meter)
        if avg:
            last_readings = meter.reading_set.filter(date__lte=end_date).order_by('-date')[:12]
            for r in last_readings:
                if is_anomaly(r.total_consumption(), avg, 2.0):
                    anomaly_count += 1
    unit = meters.first().resource_type.unit if meters.exists() else "ед."

    monthly_current = defaultdict(float)
    monthly_prev = defaultdict(float)
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        for r in readings:
            monthly_current[r.date.strftime('%Y-%m-%d')] += float(r.total_consumption())
        prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
        for r in prev_readings:
            monthly_prev[r.date.strftime('%Y-%m-%d')] += float(r.total_consumption())
    days_in_month = monthrange(year, month)[1]
    day_labels = [str(d) for d in range(1, days_in_month + 1)]
    current_data = [float(monthly_current.get(f"{year}-{month:02d}-{d:02d}", 0)) for d in range(1, days_in_month + 1)]
    prev_data = [float(monthly_prev.get(f"{year-1}-{month:02d}-{d:02d}", 0)) for d in range(1, days_in_month + 1)]

    consumption_by_type = defaultdict(float)
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            total = sum(r.total_consumption() for r in readings)
        else:
            total = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        consumption_by_type[meter.resource_type.name] += float(total)
    resource_labels = list(consumption_by_type.keys())
    resource_data = list(consumption_by_type.values())

    top_meters = []
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            total = sum(r.total_consumption() for r in readings)
        else:
            total = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        top_meters.append({'meter': meter, 'consumption': float(total), 'unit': meter.resource_type.unit})
    top_meters.sort(key=lambda x: x['consumption'], reverse=True)
    top_meters = top_meters[:5]

    anomalies = []
    for meter in meters:
        avg = get_avg_consumption(meter)
        if avg == 0:
            continue
        readings = meter.reading_set.filter(date__lte=end_date).order_by('-date')[:20]
        for r in readings:
            consumption = r.total_consumption()
            if is_anomaly(consumption, avg, 2.0):
                anomalies.append({
                    'date': r.date,
                    'meter': meter,
                    'consumption': float(consumption),
                    'avg': float(avg),
                    'ratio': float(consumption / avg) if avg else 0
                })
    anomalies.sort(key=lambda x: x['date'], reverse=True)
    anomalies = anomalies[:10]

    resource_types = ResourceType.objects.all()

    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }

    context = {
        'total_consumption': float(total_current),
        'meter_count': meter_count,
        'anomaly_count': anomaly_count,
        'unit': unit,
        'day_labels': day_labels,
        'current_data': current_data,
        'prev_data': prev_data,
        'resource_labels': resource_labels,
        'resource_data': resource_data,
        'top_meters': top_meters,
        'anomalies': anomalies,
        'year': year,
        'month': month,
        'month_name': month_names_ru[month],
        'resource_types': resource_types,
        'selected_resource_type': resource_type_id,
    }
    return context


# ------------------------------------------------------------
# ДОБАВЛЕНИЕ / РЕДАКТИРОВАНИЕ / УДАЛЕНИЕ ПОКАЗАНИЙ
# ------------------------------------------------------------
@login_required
def add_reading(request):
    if not _check_energy_view_access(request.user):
        messages.error(request, "У вас нет доступа к этому разделу.")
        return redirect('energy:meter_list')

    if request.method == 'POST':
        meter_id = request.POST.get('meter')
        date_str = request.POST.get('date')

        if not meter_id or not date_str:
            messages.error(request, "Не указан счётчик или дата.")
            return redirect('energy:meter_list')

        try:
            meter = Meter.objects.get(pk=int(meter_id))
        except Meter.DoesNotExist:
            messages.error(request, "Счётчик не найден.")
            return redirect('energy:meter_list')

        try:
            date_val = datetime.strptime(date_str, '%d.%m.%Y').date()
        except ValueError:
            try:
                date_val = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Неверный формат даты. Используйте ДД.ММ.ГГГГ или ГГГГ-ММ-ДД.")
                return redirect('energy:add_reading')

        reading = Reading(meter=meter, date=date_val)

        try:
            if meter.is_multi_tariff:
                zone_values = {}
                for key, val in request.POST.items():
                    if key.startswith('zone_'):
                        zone_values[key] = val
                reading.save()
                for key, val in zone_values.items():
                    comp_id = int(key.split('_')[1])
                    comp = TariffComponent.objects.get(pk=comp_id)
                    ZoneReading.objects.create(reading=reading, tariff_component=comp, value=Decimal(val))
                if 'document' in request.FILES:
                    reading.document = request.FILES['document']
                    reading.save(update_fields=['document'])
                meter.recalc_consumption()
            else:
                value = request.POST.get('value')
                if not value:
                    messages.error(request, "Не указано показание.")
                    return redirect('energy:add_reading')
                reading.value = Decimal(value)
                reading.save()
                if 'document' in request.FILES:
                    reading.document = request.FILES['document']
                    reading.save(update_fields=['document'])
                meter.recalc_consumption()

            log_action(user=request.user, action='CREATE', model_name='Reading', object_id=reading.pk,
                       details=f"Показание для {meter.serial_number} от {reading.date}: {reading.total_consumption()}", request=request)
            messages.success(request, "Показания успешно добавлены.")
            return redirect('energy:meter_detail', pk=meter.id)

        except ValidationError as e:
            if hasattr(e, 'error_dict'):
                for field, errors in e.error_dict.items():
                    for err in errors:
                        messages.error(request, f"{field}: {err}")
            else:
                messages.error(request, str(e))
            return redirect('energy:add_reading')

    else:
        initial = {}
        meter_id = request.GET.get('meter')
        date_str = request.GET.get('date')
        if meter_id:
            initial['meter'] = meter_id
        if date_str:
            initial['date'] = date_str
        form = ReadingForm(request.user, initial=initial)
        return render(request, 'energy/add_reading.html', {'form': form})


@login_required
def add_reading_modal(request):
    if not _check_energy_view_access(request.user):
        return JsonResponse({'status': 'error', 'message': 'Нет доступа'}, status=403)

    if request.method == 'POST':
        form = ReadingForm(request.user, request.POST, request.FILES)
        if form.is_valid():
            try:
                reading = form.save()
                log_action(user=request.user, action='CREATE', model_name='Reading', object_id=reading.pk,
                           details=f"Показание для {reading.meter.serial_number} от {reading.date}: {reading.total_consumption()}", request=request)
                return JsonResponse({'status': 'success', 'message': 'Показания добавлены'})
            except IntegrityError as e:
                meter = form.cleaned_data.get('meter')
                date_val = form.cleaned_data.get('date')
                if meter and date_val:
                    error_msg = f'Показание для счётчика "{meter.serial_number}" за дату {date_val} уже существует.'
                else:
                    error_msg = 'Показание с такими данными уже существует.'
                return JsonResponse({'status': 'error', 'errors': {'__all__': [error_msg]}}, status=400)
            except ValidationError as e:
                if hasattr(e, 'error_dict'):
                    errors = {k: [str(err) for err in v] for k, v in e.error_dict.items()}
                else:
                    errors = {'__all__': [str(e)]}
                return JsonResponse({'status': 'error', 'errors': errors}, status=400)
        else:
            return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
    else:
        meter_id = request.GET.get('meter_id')
        initial = {}
        if meter_id:
            try:
                meter = Meter.objects.get(pk=int(meter_id))
                if not can_view_meter(request.user, meter):
                    return JsonResponse({'status': 'error', 'message': 'Нет доступа'}, status=403)
                initial['meter'] = meter.id
            except Meter.DoesNotExist:
                pass
        form = ReadingForm(request.user, initial=initial)
        html = render_to_string('energy/add_reading_modal_form.html', {'form': form}, request=request)
        return JsonResponse({'html': html})


@login_required
def edit_reading(request, pk):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на редактирование показаний.")
        return redirect('energy:meter_list')

    reading = get_object_or_404(Reading, pk=pk)
    if not can_edit_reading(request.user, reading):
        messages.error(request, "У вас нет прав на редактирование этого показания.")
        return redirect('energy:meter_list')

    if request.method == 'POST':
        form = ReadingEditForm(request.user, reading, request.POST, request.FILES)
        if form.is_valid():
            form.save()

            avg = get_avg_consumption(reading.meter)
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg):
                messages.warning(request, f'⚠️ Внимание! Потребление за {reading.date} ({consumption:.2f} {reading.meter.resource_type.unit}) значительно превышает среднее ({avg:.2f} {reading.meter.resource_type.unit})!')
            log_action(user=request.user, action='EDIT', model_name='Reading', object_id=reading.pk,
                       details=f"Показание для {reading.meter.serial_number} от {reading.date}: {reading.total_consumption()}", request=request)
            messages.success(request, 'Показания обновлены.')
            next_url = request.POST.get('next', reverse('energy:meter_detail', args=[reading.meter.id]))
            return redirect(next_url)
    else:
        form = ReadingEditForm(request.user, reading)

    return render(request, 'energy/reading_edit.html', {'form': form, 'reading': reading})


@login_required
def delete_reading(request, pk):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на удаление показаний.")
        return redirect('energy:meter_list')

    reading = get_object_or_404(Reading, pk=pk)
    if not can_delete_reading(request.user, reading):
        messages.error(request, "У вас нет прав на удаление этого показания.")
        return redirect('energy:meter_list')

    if request.method == 'POST':
        next_url = request.POST.get('next', reverse('energy:meter_detail', args=[reading.meter.id]))
        serial = reading.meter.serial_number
        if reading.document:
            try:
                reading.document.delete(save=False)
            except Exception:
                pass
        reading.delete()
        log_action(user=request.user, action='DELETE', model_name='Reading', object_id=pk,
                   details=f"Удалено показание для {serial} от {reading.date}", request=request)
        messages.success(request, f'Показание счётчика "{serial}" удалено.')
        return redirect(next_url)

    return render(request, 'energy/reading_confirm_delete.html', {'reading': reading})


# ------------------------------------------------------------
# ЗАЩИЩЁННЫЙ ПРОСМОТР ФАЙЛА
# ------------------------------------------------------------
@login_required
def reading_document_view(request, reading_id):
    reading = get_object_or_404(Reading, pk=reading_id)
    if not can_view_meter(request.user, reading.meter):
        raise PermissionDenied('У вас нет доступа к этому файлу')
    if not reading.document or not reading.document.name:
        raise Http404('Файл не найден')
    return redirect(reading.document.url)


# ------------------------------------------------------------
# ДОБАВЛЕНИЕ / РЕДАКТИРОВАНИЕ / УДАЛЕНИЕ СЧЁТЧИКОВ
# ------------------------------------------------------------
@login_required
def add_meter(request):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на добавление счётчиков.")
        return redirect('energy:meter_list')
    if request.method == 'POST':
        form = MeterForm(request.user, request.POST)
        if form.is_valid():
            meter = form.save()
            log_action(user=request.user, action='CREATE', model_name='Meter', object_id=meter.pk,
                       details=f"Добавлен счётчик: {meter.serial_number}, тип: {meter.resource_type.name}", request=request)
            messages.success(request, f'Счётчик "{meter.serial_number}" успешно добавлен.')
            return redirect('energy:meter_list')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = MeterForm(request.user)
    return render(request, 'energy/add_meter.html', {'form': form})


class MeterListView(LoginRequiredMixin, ListView):
    model = Meter
    template_name = 'energy/meter_list.html'
    context_object_name = 'meters'
    paginate_by = 12

    def get_queryset(self):
        queryset = Meter.objects.all()

        # --- Быстрые фильтры (чипсы) ---
        # Фильтр по типу (технический / основной)
        meter_type = self.request.GET.get('meter_type')
        if meter_type == 'technical':
            queryset = queryset.filter(is_technical=True)
        elif meter_type == 'main':
            queryset = queryset.filter(is_technical=False)

        # Фильтр по активности
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        # Фильтр по типу ресурса (чипсы)
        resource_type_chip = self.request.GET.get('resource_type_chip')
        if resource_type_chip:
            queryset = queryset.filter(resource_type_id=resource_type_chip)

        # --- Поиск ---
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(serial_number__icontains=search) |
                models.Q(location__icontains=search) |
                models.Q(resource_type__name__icontains=search)
            )

        # --- Расширенный фильтр ---
        # Тип ресурса (выпадающий список)
        resource_type = self.request.GET.get('resource_type')
        if resource_type:
            queryset = queryset.filter(resource_type_id=resource_type)

        # Технический учёт (выпадающий список)
        is_technical = self.request.GET.get('is_technical')
        if is_technical == '1':
            queryset = queryset.filter(is_technical=True)
        elif is_technical == '0':
            queryset = queryset.filter(is_technical=False)

        # Активность (выпадающий список)
        is_active = self.request.GET.get('is_active')
        if is_active == '1':
            queryset = queryset.filter(is_active=True)
        elif is_active == '0':
            queryset = queryset.filter(is_active=False)

        # Диапазон суммы потребления (за всё время) – можно использовать агрегацию, но для простоты оставим пока заглушку
        # Здесь можно добавить фильтрацию по total_consumption через аннотацию, но это может быть сложно.
        # Пока оставим как есть, но поля в форме добавим.

        # Диапазон дат последнего показания – можно через subquery, но для простоты пока пропустим.

        return queryset.order_by('serial_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_edit'] = can_edit_all_meters(self.request.user)
        context['can_delete'] = can_delete_meter(self.request.user, None)

        # Статистика
        all_meters = Meter.objects.all()
        context['total_count'] = all_meters.count()
        context['active_count'] = all_meters.filter(is_active=True).count()
        context['inactive_count'] = all_meters.filter(is_active=False).count()
        context['technical_count'] = all_meters.filter(is_technical=True).count()
        context['main_count'] = all_meters.filter(is_technical=False).count()

        # Типы ресурсов для фильтров
        context['resource_types'] = ResourceType.objects.all()

        # Текущие параметры для сохранения в форме
        context['filter_data'] = {
            'search': self.request.GET.get('search', ''),
            'resource_type': self.request.GET.get('resource_type', ''),
            'is_technical': self.request.GET.get('is_technical', ''),
            'is_active': self.request.GET.get('is_active', ''),
            'meter_type': self.request.GET.get('meter_type', ''),
            'status': self.request.GET.get('status', ''),
            'resource_type_chip': self.request.GET.get('resource_type_chip', ''),
        }

        # Для чипсов – список уникальных типов ресурсов, используемых в приборах
        used_resource_types = ResourceType.objects.filter(meter__isnull=False).distinct()
        context['used_resource_types'] = used_resource_types

        return context


class MeterDetailView(LoginRequiredMixin, DetailView):
    model = Meter
    template_name = 'energy/meter_detail.html'
    context_object_name = 'meter'

    def dispatch(self, request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or not can_view_meter(request.user, self.get_object()):
            messages.error(request, "У вас нет доступа к деталям счётчика.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Meter.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        meter = self.object

        context['can_edit'] = can_edit_meter(self.request.user, meter)
        context['can_delete'] = can_delete_meter(self.request.user, meter)
        log_action(user=self.request.user, action='VIEW', model_name='Meter', object_id=meter.pk,
                   details=f"Просмотр счётчика {meter.serial_number}", request=self.request)

        months_count = int(self.request.GET.get('months', 12))
        months_count = min(months_count, 36)
        chart_type = self.request.GET.get('chart_type', 'line')
        context['selected_months'] = months_count
        context['chart_type'] = chart_type

        today = date.today()
        start_date = today.replace(day=1) - timedelta(days=1)
        for _ in range(months_count - 1):
            start_date = start_date.replace(day=1) - timedelta(days=1)
        start_date = start_date.replace(day=1)

        readings = meter.reading_set.filter(date__gte=start_date).order_by('date')

        if meter.is_multi_tariff:
            zones = TariffComponent.objects.filter(
                resource_type=meter.resource_type,
                is_multi_tariff_zone=True
            ).order_by('name')
            zone_names = [zone.name for zone in zones]
            monthly_data = defaultdict(lambda: defaultdict(float))
            for reading in readings:
                month_key = reading.date.strftime('%Y-%m')
                for zr in reading.zone_readings.all():
                    if zr.consumption:
                        monthly_data[month_key][zr.tariff_component.name] += float(zr.consumption)
            months = sorted(monthly_data.keys())
            zone_datasets = []
            fixed_colors = [
                {'bg': 'rgba(13, 110, 253, 0.6)', 'border': 'rgba(13, 110, 253, 1)'},
                {'bg': 'rgba(220, 53, 69, 0.6)', 'border': 'rgba(220, 53, 69, 1)'},
                {'bg': 'rgba(25, 135, 84, 0.6)', 'border': 'rgba(25, 135, 84, 1)'},
                {'bg': 'rgba(255, 193, 7, 0.6)', 'border': 'rgba(255, 193, 7, 1)'},
                {'bg': 'rgba(111, 66, 193, 0.6)', 'border': 'rgba(111, 66, 193, 1)'},
                {'bg': 'rgba(23, 162, 184, 0.6)', 'border': 'rgba(23, 162, 184, 1)'},
            ]
            for idx, zone in enumerate(zone_names):
                data = [monthly_data[month].get(zone, 0) for month in months]
                color = fixed_colors[idx % len(fixed_colors)]
                zone_datasets.append({
                    'label': zone,
                    'data': data,
                    'backgroundColor': color['bg'],
                    'borderColor': color['border'],
                    'borderWidth': 1,
                    'stack': 'zones',
                })
            context['month_labels'] = months
            context['zone_datasets'] = zone_datasets
            context['is_multi_tariff'] = True
        else:
            monthly_consumption = defaultdict(float)
            for reading in readings:
                if reading.consumption:
                    month_key = reading.date.strftime('%Y-%m')
                    monthly_consumption[month_key] += float(reading.consumption)
            months = sorted(monthly_consumption.keys())
            consumptions = [monthly_consumption[month] for month in months]
            context['month_labels'] = months
            context['consumptions'] = consumptions
            context['is_multi_tariff'] = False

        last_readings = meter.reading_set.order_by('-date')[:5]
        last_readings_with_anomaly = []
        if last_readings:
            avg = get_avg_consumption(meter)
            for reading in last_readings:
                consumption = reading.total_consumption()
                anomaly = is_anomaly(consumption, avg)
                last_readings_with_anomaly.append({'reading': reading, 'anomaly': anomaly})
        context['last_readings_with_anomaly'] = last_readings_with_anomaly

        documents = meter.documents.all().order_by('-uploaded_at')
        paginator = Paginator(documents, 15)
        page = self.request.GET.get('doc_page')
        try:
            docs_page = paginator.page(page)
        except PageNotAnInteger:
            docs_page = paginator.page(1)
        except EmptyPage:
            docs_page = paginator.page(paginator.num_pages)
        context['documents_page'] = docs_page
        context['base_url'] = reverse('energy:meter_detail', args=[meter.pk])

        return context


class MeterUpdateView(LoginRequiredMixin, UpdateView):
    model = Meter
    form_class = MeterForm
    template_name = 'energy/meter_edit.html'
    success_url = reverse_lazy('energy:meter_list')

    def dispatch(self, request, *args, **kwargs):
        if not _check_energy_edit_access(request.user):
            messages.error(request, "У вас нет прав на редактирование счётчиков.")
            return redirect('energy:meter_list')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        if can_edit_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        log_action(user=self.request.user, action='EDIT', model_name='Meter', object_id=self.object.pk,
                   details=f"Редактирование счётчика {self.object.serial_number}", request=self.request)
        return response


class MeterDeleteView(LoginRequiredMixin, DeleteView):
    model = Meter
    template_name = 'energy/meter_confirm_delete.html'
    success_url = reverse_lazy('energy:meter_list')

    def dispatch(self, request, *args, **kwargs):
        if not _check_energy_edit_access(request.user):
            messages.error(request, "У вас нет прав на удаление счётчиков.")
            return redirect('energy:meter_list')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        if can_edit_all_meters(self.request.user):
            return Meter.objects.all()
        return Meter.objects.all()

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        serial = obj.serial_number
        log_action(user=request.user, action='DELETE', model_name='Meter', object_id=obj.pk,
                   details=f"Удаление счётчика {serial}", request=request)
        return super().delete(request, *args, **kwargs)


# ------------------------------------------------------------
# ДОКУМЕНТЫ СЧЁТЧИКА
# ------------------------------------------------------------
@login_required
def upload_document(request, meter_pk):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на загрузку документов.")
        return redirect('energy:meter_list')
    meter = get_object_or_404(Meter, pk=meter_pk)
    if not can_edit_meter(request.user, meter):
        messages.error(request, "У вас нет прав на загрузку документов для этого счётчика.")
        return redirect('energy:meter_detail', pk=meter.pk)
    if request.method == 'POST':
        form = MeterDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.meter = meter
            doc.save()
            log_action(user=request.user, action='CREATE', model_name='MeterDocument', object_id=doc.pk,
                       details=f"Загружен документ для счётчика {meter.serial_number}: {doc.get_file_name()}", request=request)
            messages.success(request, f'Документ "{doc.get_file_name()}" загружен.')
            return redirect('energy:meter_detail', pk=meter.pk)
    else:
        form = MeterDocumentForm()
    return render(request, 'energy/upload_document.html', {'form': form, 'meter': meter})


@login_required
def delete_document(request, pk):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на удаление документов.")
        return redirect('energy:meter_list')
    doc = get_object_or_404(MeterDocument, pk=pk)
    if not can_edit_meter(request.user, doc.meter):
        messages.error(request, "У вас нет прав на удаление этого документа.")
        return redirect('energy:meter_detail', pk=doc.meter.pk)
    meter_pk = doc.meter.pk
    file_name = doc.get_file_name()
    doc.delete()
    log_action(user=request.user, action='DELETE', model_name='MeterDocument', object_id=pk,
               details=f"Удалён документ {file_name} для счётчика {doc.meter.serial_number}", request=request)
    messages.success(request, 'Документ удалён.')
    return redirect('energy:meter_detail', pk=meter_pk)


# ------------------------------------------------------------
# ОТЧЁТЫ
# ------------------------------------------------------------
@login_required
def consumption_report(request):
    if not _check_energy_view_access(request.user):
        messages.error(request, "У вас нет доступа к отчёту.")
        return redirect('home')

    period_type = request.GET.get('period', 'month')
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    quarter = int(request.GET.get('quarter', (datetime.now().month - 1) // 3 + 1))
    group_id = request.GET.get('group', 'all')
    chart_type = request.GET.get('chart_type', 'bar')
    meter_type = request.GET.get('meter_type', 'main')  # 'main', 'technical', 'all'

    month_names_ru = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь',
                      7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}

    if period_type == 'month':
        start_date = date(year, month, 1)
        end_date = start_date.replace(day=monthrange(year, month)[1])
        period_label = f"{month_names_ru[month]} {year}"
        prev_start_date = date(year-1, month, 1)
        prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    elif period_type == 'quarter':
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
        period_label = f"{quarter} квартал {year}"
        prev_start_date = date(year-1, start_month, 1)
        prev_end_date = prev_start_date + relativedelta(months=3) - timedelta(days=1)
        date_range = [date(year, m, 1) for m in range(start_month, start_month+3)]
    else:  # year
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = str(year)
        prev_start_date = date(year-1, 1, 1)
        prev_end_date = date(year-1, 12, 31)
        date_range = [date(year, m, 1) for m in range(1, 13)]

    # Фильтр по типу приборов
    meters = Meter.objects.filter(is_active=True)
    if meter_type == 'main':
        meters = meters.filter(is_technical=False)
    elif meter_type == 'technical':
        meters = meters.filter(is_technical=True)
    # если 'all' – оставляем все

    if group_id != 'all':
        meters = meters.filter(resource_type_id=group_id)

    # ---- Данные для таблицы ----
    table_data = []
    total_current = Decimal('0')
    total_prev = Decimal('0')
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            current = sum(r.total_consumption() for r in readings)
        else:
            current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        if current == 0:
            continue
        prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
        if prev_readings:
            if meter.is_multi_tariff:
                prev = sum(r.total_consumption() for r in prev_readings)
            else:
                prev = prev_readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        else:
            prev = Decimal('0')
        if prev != 0:
            change = (current - prev) / prev * 100
            changetype = 'increase' if change > 0 else 'decrease' if change < 0 else 'stable'
        else:
            change = 100 if current > 0 else 0
            changetype = 'increase' if current > 0 else 'stable'
        total_current += current
        total_prev += prev
        table_data.append({
            'meter': meter,
            'consumption': current,
            'prev_consumption': prev,
            'change_percent': change,
            'changetype': changetype
        })
    table_data.sort(key=lambda x: x['consumption'], reverse=True)

    # ---- Данные для графика динамики ----
    chart_labels = []
    chart_current = []
    chart_prev = []
    if period_type == 'month':
        for d in date_range:
            chart_labels.append(d.strftime('%d.%m'))
            day_readings = Reading.objects.filter(
                meter__in=meters,
                date=d
            )
            total_day = sum(r.total_consumption() for r in day_readings)
            chart_current.append(float(total_day))
            prev_d = d.replace(year=d.year-1)
            prev_day_readings = Reading.objects.filter(
                meter__in=meters,
                date=prev_d
            )
            total_prev_day = sum(r.total_consumption() for r in prev_day_readings)
            chart_prev.append(float(total_prev_day))
    else:
        for d in date_range:
            chart_labels.append(month_names_ru[d.month])
            month_readings = Reading.objects.filter(
                meter__in=meters,
                date__year=d.year,
                date__month=d.month
            )
            total_month = sum(r.total_consumption() for r in month_readings)
            chart_current.append(float(total_month))
            prev_d = d.replace(year=d.year-1)
            prev_month_readings = Reading.objects.filter(
                meter__in=meters,
                date__year=prev_d.year,
                date__month=prev_d.month
            )
            total_prev_month = sum(r.total_consumption() for r in prev_month_readings)
            chart_prev.append(float(total_prev_month))

    # ---- Данные для круговой диаграммы по типам ресурсов ----
    consumption_by_type = defaultdict(float)
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            total = sum(r.total_consumption() for r in readings)
        else:
            total = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        consumption_by_type[meter.resource_type.name] += float(total)
    pie_labels = list(consumption_by_type.keys())
    pie_data = list(consumption_by_type.values())

    # ---- Общие итоги ----
    units = {item['meter'].resource_type.unit for item in table_data}
    total_unit = units.pop() if len(units) == 1 else "ед."

    if total_prev != 0:
        total_change = (total_current - total_prev) / total_prev * 100
        total_changetype = 'increase' if total_change > 0 else 'decrease' if total_change < 0 else 'stable'
    else:
        total_change = 100 if total_current > 0 else 0
        total_changetype = 'increase' if total_current > 0 else 'stable'

    current_year = datetime.now().year
    years = range(current_year - 2, current_year + 1)
    resource_types = ResourceType.objects.all()

    context = {
        'table_data': table_data,
        'total_current': total_current,
        'total_prev': total_prev,
        'total_change': total_change,
        'total_changetype': total_changetype,
        'total_unit': total_unit,
        'period_type': period_type,
        'period_label': period_label,
        'selected_year': year,
        'selected_month': month,
        'selected_quarter': quarter,
        'selected_group': group_id,
        'years': years,
        'months': [(1, 'Январь'), (2, 'Февраль'), (3, 'Март'), (4, 'Апрель'),
                   (5, 'Май'), (6, 'Июнь'), (7, 'Июль'), (8, 'Август'),
                   (9, 'Сентябрь'), (10, 'Октябрь'), (11, 'Ноябрь'), (12, 'Декабрь')],
        'quarters': [1, 2, 3, 4],
        'resource_types': resource_types,
        'chart_labels': chart_labels,
        'chart_current': chart_current,
        'chart_prev': chart_prev,
        'chart_type': chart_type,
        'pie_labels': pie_labels,
        'pie_data': pie_data,
        'meter_type': meter_type,
    }
    return render(request, 'energy/consumption_report.html', context)


@login_required
def export_consumption_report(request):
    if not _check_energy_view_access(request.user):
        messages.error(request, "У вас нет доступа к экспорту отчёта.")
        return redirect('home')
    period_type = request.GET.get('period', 'month')
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    quarter = int(request.GET.get('quarter', (datetime.now().month - 1) // 3 + 1))
    group_id = request.GET.get('group', 'all')
    month_names_ru = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
    if period_type == 'month':
        start_date = date(year, month, 1)
        end_date = start_date.replace(day=monthrange(year, month)[1])
        period_label = f"{month_names_ru[month]} {year}"
        prev_start_date = date(year-1, month, 1)
        prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])
    elif period_type == 'quarter':
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
        period_label = f"{quarter} квартал {year}"
        prev_start_date = date(year-1, start_month, 1)
        prev_end_date = prev_start_date + relativedelta(months=3) - timedelta(days=1)
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = str(year)
        prev_start_date = date(year-1, 1, 1)
        prev_end_date = date(year-1, 12, 31)

    meters = Meter.objects.filter(is_active=True, is_technical=False)
    if group_id != 'all':
        meters = meters.filter(resource_type_id=group_id)

    table_data = []
    total_current = Decimal('0')
    total_prev = Decimal('0')
    for meter in meters:
        readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
        if not readings:
            continue
        if meter.is_multi_tariff:
            current = sum(r.total_consumption() for r in readings)
        else:
            current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        if current == 0:
            continue
        prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
        if prev_readings:
            if meter.is_multi_tariff:
                prev = sum(r.total_consumption() for r in prev_readings)
            else:
                prev = prev_readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
        else:
            prev = Decimal('0')
        if prev != 0:
            change = (current - prev) / prev * 100
        else:
            change = 100 if current > 0 else 0
        total_current += current
        total_prev += prev
        table_data.append({'meter': meter, 'consumption': current, 'prev_consumption': prev, 'change_percent': change})
    table_data.sort(key=lambda x: x['consumption'], reverse=True)

    if total_prev != 0:
        total_change = (total_current - total_prev) / total_prev * 100
    else:
        total_change = 100 if total_current > 0 else 0

    log_action(user=request.user, action='EXPORT', model_name='Report', object_id='',
               details=f"Экспорт отчёта по потреблению в Excel (период: {period_label}, фильтр: {group_id})", request=request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт {period_label}"
    headers = ['№', 'Прибор учёта (серийный номер)', 'Тип ресурса', 'Тип тарифа',
               f'Потребление за {period_label} (ед.)', f'Потребление за аналогичный период прошлого года (ед.)', 'Динамика, %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row, item in enumerate(table_data, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=item['meter'].serial_number)
        ws.cell(row=row, column=3, value=f"{item['meter'].resource_type.name} ({item['meter'].resource_type.unit})")
        ws.cell(row=row, column=4, value="Многотарифный" if item['meter'].is_multi_tariff else "Одноставочный")
        ws.cell(row=row, column=5, value=float(item['consumption']))
        ws.cell(row=row, column=6, value=float(item['prev_consumption']))
        ws.cell(row=row, column=7, value=float(item['change_percent']))
    last_row = len(table_data) + 2
    ws.cell(row=last_row, column=4, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=float(total_current)).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=float(total_prev)).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=float(total_change)).font = Font(bold=True)
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="consumption_report_{period_label}.xlsx"'
    wb.save(response)
    return response


@login_required
def anomaly_report(request):
    if not _check_energy_view_access(request.user):
        messages.error(request, "У вас нет доступа к отчёту по аномалиям.")
        return redirect('home')
    resource_type_id = request.GET.get('resource_type')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    threshold = float(request.GET.get('threshold', 2.0))
    meters = Meter.objects.filter(is_active=True, is_technical=False)
    if resource_type_id and resource_type_id != 'all':
        meters = meters.filter(resource_type_id=resource_type_id)
    anomalies = []
    for meter in meters:
        avg = get_avg_consumption(meter)
        if avg == 0:
            continue
        readings = meter.reading_set.order_by('-date')
        if start_date_str:
            readings = readings.filter(date__gte=start_date_str)
        if end_date_str:
            readings = readings.filter(date__lte=end_date_str)
        for reading in readings:
            consumption = reading.total_consumption()
            if is_anomaly(consumption, avg, threshold):
                anomalies.append({'date': reading.date, 'meter': meter, 'consumption': consumption, 'avg': avg, 'ratio': float(consumption / avg) if avg else 0})
    resource_types = ResourceType.objects.all()
    log_action(user=request.user, action='VIEW', model_name='Report', object_id='',
               details=f"Просмотр отчёта по аномалиям (порог {threshold})", request=request)
    context = {'anomalies': anomalies, 'resource_types': resource_types, 'selected_resource_type': resource_type_id,
               'start_date': start_date_str, 'end_date': end_date_str, 'threshold': threshold}
    return render(request, 'energy/anomaly_report.html', context)


# ------------------------------------------------------------
# ИМПОРТ / ЭКСПОРТ
# ------------------------------------------------------------
@login_required
def import_readings(request):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на импорт показаний.")
        return redirect('energy:meter_list')

    if request.method == 'POST':
        form = ImportReadingsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            dry_run = form.cleaned_data.get('dry_run', False)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

            try:
                meter_col = headers.index('meter_serial') + 1
            except ValueError:
                meter_col = None
            try:
                date_col = headers.index('date') + 1
            except ValueError:
                date_col = None
            try:
                value_col = headers.index('value') + 1
            except ValueError:
                value_col = None

            zone_columns = [(i+1, h) for i, h in enumerate(headers)
                            if h and h not in ['meter_serial', 'date', 'value']
                            and TariffComponent.objects.filter(name=h, is_multi_tariff_zone=True).exists()]

            if not meter_col or not date_col:
                messages.error(request, "Файл должен содержать колонки 'meter_serial' и 'date'")
                return redirect('energy:import_readings')

            if not value_col and not zone_columns:
                messages.error(request, "Файл должен содержать колонку 'value' для однотарифных или колонки с названиями зон для многотарифных счётчиков")
                return redirect('energy:import_readings')

            created_count = 0
            updated_count = 0
            errors = []
            affected_meters = set()

            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or all(cell is None for cell in row):
                        continue

                    meter_serial = row[meter_col-1] if meter_col and len(row) >= meter_col else None
                    date_str_val = row[date_col-1] if date_col and len(row) >= date_col else None

                    if not meter_serial or not date_str_val:
                        errors.append(f"Строка {row_idx}: отсутствует серийный номер или дата")
                        continue

                    try:
                        meter = Meter.objects.get(serial_number=str(meter_serial))
                    except Meter.DoesNotExist:
                        errors.append(f"Строка {row_idx}: счётчик с серийным номером '{meter_serial}' не найден")
                        continue

                    try:
                        if isinstance(date_str_val, datetime):
                            reading_date = date_str_val.date()
                        else:
                            reading_date = date_parser.parse(str(date_str_val)).date()
                    except Exception:
                        errors.append(f"Строка {row_idx}: неверный формат даты (ожидается YYYY-MM-DD или подобный)")
                        continue

                    if reading_date > date.today():
                        errors.append(f"Строка {row_idx}: дата {reading_date} не может быть в будущем")
                        continue

                    if meter.is_multi_tariff:
                        if not zone_columns:
                            errors.append(f"Строка {row_idx}: счётчик многотарифный, но в файле нет колонок для зон")
                            continue

                        zone_values = {}
                        valid = True
                        for col_num, zone_name in zone_columns:
                            val = row[col_num-1] if len(row) >= col_num else None
                            if val is None:
                                errors.append(f"Строка {row_idx}: отсутствует значение для зоны '{zone_name}'")
                                valid = False
                                break
                            try:
                                zone_values[zone_name] = Decimal(str(val))
                            except Exception:
                                errors.append(f"Строка {row_idx}: неверное числовое значение для зоны '{zone_name}'")
                                valid = False
                                break
                        if not valid:
                            continue

                        components = {comp.name: comp for comp in TariffComponent.objects.filter(
                            resource_type=meter.resource_type,
                            is_multi_tariff_zone=True,
                            valid_from__lte=reading_date
                        ).exclude(valid_to__lt=reading_date)}

                        missing_zones = set(zone_values.keys()) - set(components.keys())
                        if missing_zones:
                            errors.append(f"Строка {row_idx}: неизвестные зоны: {', '.join(missing_zones)}")
                            continue

                        if dry_run:
                            created_count += 1
                            continue

                        reading, created = Reading.objects.update_or_create(
                            meter=meter,
                            date=reading_date,
                            defaults={'value': None}
                        )
                        for zone_name, val in zone_values.items():
                            comp = components[zone_name]
                            ZoneReading.objects.update_or_create(
                                reading=reading,
                                tariff_component=comp,
                                defaults={'value': val}
                            )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        affected_meters.add(meter.pk)
                        log_action(user=request.user, action='IMPORT', model_name='Reading', object_id=reading.pk,
                                   details=f"Импорт показаний для {meter.serial_number} от {reading_date}", request=request)

                    else:
                        if not value_col:
                            errors.append(f"Строка {row_idx}: для однотарифного счётчика нужна колонка 'value'")
                            continue

                        val = row[value_col-1] if len(row) >= value_col else None
                        if val is None:
                            errors.append(f"Строка {row_idx}: отсутствует значение показания")
                            continue
                        try:
                            value = Decimal(str(val))
                        except Exception:
                            errors.append(f"Строка {row_idx}: неверное числовое значение для показания")
                            continue

                        if dry_run:
                            created_count += 1
                            continue

                        reading, created = Reading.objects.update_or_create(
                            meter=meter,
                            date=reading_date,
                            defaults={'value': value}
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1
                        affected_meters.add(meter.pk)
                        log_action(user=request.user, action='IMPORT', model_name='Reading', object_id=reading.pk,
                                   details=f"Импорт показаний для {meter.serial_number} от {reading_date}", request=request)

            if not dry_run:
                for meter_pk in affected_meters:
                    meter = Meter.objects.get(pk=meter_pk)
                    meter.recalc_consumption()

            if errors:
                messages.warning(request, f"Импорт завершён с ошибками. Создано/обновлено: {created_count}, пропущено строк: {len(errors)}. Ошибки: {', '.join(errors[:10])}")
            else:
                messages.success(request, f"Импорт успешно завершён. Создано: {created_count}, обновлено: {updated_count}")

            if dry_run:
                messages.info(request, "Режим проверки (dry run) – данные не сохранены")

            return redirect('energy:import_readings')
    else:
        form = ImportReadingsForm()

    return render(request, 'energy/import_readings.html', {'form': form})


@login_required
def download_import_template(request):
    if not _check_energy_edit_access(request.user):
        messages.error(request, "У вас нет прав на скачивание шаблона.")
        return redirect('energy:meter_list')
    log_action(user=request.user, action='EXPORT', model_name='Template', object_id='', details="Скачивание шаблона импорта показаний", request=request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Импорт показаний"
    headers = ['meter_serial', 'date', 'value']
    example_zones = TariffComponent.objects.filter(is_multi_tariff_zone=True).values_list('name', flat=True)[:2]
    for zone in example_zones:
        headers.append(zone)
    ws.append(headers)
    meter = Meter.objects.filter(is_active=True).first()
    if meter:
        ws.append([meter.serial_number, '2026-01-15', '123.45'] + [''] * len(example_zones))
        ws.append([meter.serial_number, '2026-02-15', '234.56'] + [''] * len(example_zones))
    else:
        ws.append(['СЧ-001', '2026-01-15', '123.45'] + [''] * len(example_zones))
        ws.append(['СЧ-001', '2026-02-15', '234.56'] + [''] * len(example_zones))
    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="import_readings_template.xlsx"'
    wb.save(response)
    return response


@login_required
def export_readings(request):
    if not _check_energy_view_access(request.user):
        messages.error(request, "У вас нет доступа к экспорту показаний.")
        return redirect('energy:meter_list')
    meter_id = request.GET.get('meter')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if not meter_id:
        meters = Meter.objects.all()
        if not can_view_all_meters(request.user):
            meters = meters.filter(user=request.user)
        return render(request, 'energy/export_readings.html', {'meters': meters})
    meter = get_object_or_404(Meter, pk=meter_id)
    if not can_view_meter(request.user, meter):
        messages.error(request, "Нет доступа к счётчику")
        return redirect('energy:meter_list')
    readings = meter.reading_set.all().order_by('-date')
    if start_date:
        readings = readings.filter(date__gte=start_date)
    if end_date:
        readings = readings.filter(date__lte=end_date)
    if not readings.exists():
        messages.warning(request, "Нет показаний за выбранный период")
        return redirect('energy:export_readings')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История показаний"
    if meter.is_multi_tariff:
        zones = TariffComponent.objects.filter(resource_type=meter.resource_type, is_multi_tariff_zone=True).order_by('name')
        headers = ['Дата'] + [zone.name for zone in zones] + ['Суммарное потребление (ед.)']
        ws.append(headers)
        for reading in readings:
            row = [reading.date.strftime('%d.%m.%Y')]
            zone_values = {zr.tariff_component.name: zr.value for zr in reading.zone_readings.all()}
            for zone in zones:
                row.append(zone_values.get(zone.name, ''))
            row.append(float(reading.total_consumption()))
            ws.append(row)
    else:
        headers = ['Дата', 'Показание (суммарное)', 'Потребление (ед.)']
        ws.append(headers)
        for reading in readings:
            ws.append([reading.date.strftime('%d.%m.%Y'), float(reading.value) if reading.value else '', float(reading.consumption) if reading.consumption else ''])
    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"readings_{meter.serial_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ------------------------------------------------------------
# НАСТРАИВАЕМЫЙ ДАШБОРД
# ------------------------------------------------------------
AVAILABLE_WIDGETS = {
    'kpi': {'title': 'Ключевые показатели', 'order': 1},
    'consumption_chart': {'title': 'Динамика потребления', 'order': 2},
    'resource_pie': {'title': 'Распределение по типам ресурсов', 'order': 3},
    'top_meters': {'title': 'Топ-5 счётчиков по потреблению', 'order': 4},
    'anomaly_table': {'title': 'Последние аномалии', 'order': 5},
    'recent_readings': {'title': 'Последние показания', 'order': 6},
}

def get_widget_data(widget_name, request, filters):
    from django.db.models import Sum
    period_type = filters.get('period', 'month')
    year = filters.get('year', date.today().year)
    month = filters.get('month', date.today().month)
    quarter = filters.get('quarter', (date.today().month - 1) // 3 + 1)
    resource_type_id = filters.get('resource_type', 'all')
    anomaly_threshold = filters.get('threshold', 2.0)
    if period_type == 'month':
        start_date = date(year, month, 1)
        end_date = start_date.replace(day=monthrange(year, month)[1])
        prev_start_date = date(year-1, month, 1)
        prev_end_date = prev_start_date.replace(day=monthrange(year-1, month)[1])
    elif period_type == 'quarter':
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1)
        end_date = start_date + relativedelta(months=3) - timedelta(days=1)
        prev_start_date = date(year-1, start_month, 1)
        prev_end_date = prev_start_date + relativedelta(months=3) - timedelta(days=1)
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        prev_start_date = date(year-1, 1, 1)
        prev_end_date = date(year-1, 12, 31)
    meters = Meter.objects.filter(is_active=True, is_technical=False)
    if resource_type_id != 'all':
        meters = meters.filter(resource_type_id=resource_type_id)
    if widget_name == 'kpi':
        total_current = Decimal('0')
        meter_count = meters.count()
        anomaly_count = 0
        for meter in meters:
            readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
            if not readings: continue
            if meter.is_multi_tariff:
                current = sum(r.total_consumption() for r in readings)
            else:
                current = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
            total_current += current
            avg = get_avg_consumption(meter)
            if avg:
                last_readings = meter.reading_set.filter(date__lte=end_date).order_by('-date')[:12]
                for r in last_readings:
                    if is_anomaly(r.total_consumption(), avg, anomaly_threshold):
                        anomaly_count += 1
        unit = meters.first().resource_type.unit if meters.exists() else "ед."
        return {'total_consumption': float(total_current), 'meter_count': meter_count, 'anomaly_count': anomaly_count, 'unit': unit}
    elif widget_name == 'consumption_chart':
        monthly_current = defaultdict(float)
        monthly_prev = defaultdict(float)
        for meter in meters:
            readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
            for r in readings:
                monthly_current[r.date.strftime('%Y-%m')] += float(r.total_consumption())
            prev_readings = meter.reading_set.filter(date__gte=prev_start_date, date__lte=prev_end_date)
            for r in prev_readings:
                monthly_prev[r.date.strftime('%Y-%m')] += float(r.total_consumption())
        months = sorted(set(monthly_current.keys()) | set(monthly_prev.keys()))
        return {'labels': months, 'current': [monthly_current[m] for m in months], 'previous': [monthly_prev[m] for m in months], 'unit': meters.first().resource_type.unit if meters.exists() else "ед."}
    elif widget_name == 'resource_pie':
        consumption_by_type = defaultdict(float)
        for meter in meters:
            readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
            if not readings: continue
            if meter.is_multi_tariff:
                total = sum(r.total_consumption() for r in readings)
            else:
                total = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
            consumption_by_type[meter.resource_type.name] += float(total)
        return {'labels': list(consumption_by_type.keys()), 'data': list(consumption_by_type.values())}
    elif widget_name == 'top_meters':
        top = []
        for meter in meters:
            readings = meter.reading_set.filter(date__gte=start_date, date__lte=end_date)
            if not readings: continue
            if meter.is_multi_tariff:
                total = sum(r.total_consumption() for r in readings)
            else:
                total = readings.aggregate(total=Sum('consumption'))['total'] or Decimal('0')
            top.append({'meter': meter, 'consumption': float(total), 'unit': meter.resource_type.unit})
        top.sort(key=lambda x: x['consumption'], reverse=True)
        return top[:5]
    elif widget_name == 'anomaly_table':
        anomalies = []
        for meter in meters:
            avg = get_avg_consumption(meter)
            if avg == 0: continue
            readings = meter.reading_set.filter(date__lte=end_date).order_by('-date')[:20]
            for r in readings:
                consumption = r.total_consumption()
                if is_anomaly(consumption, avg, anomaly_threshold):
                    anomalies.append({'date': r.date, 'meter': meter, 'consumption': float(consumption), 'avg': float(avg), 'ratio': float(consumption / avg) if avg else 0})
        anomalies.sort(key=lambda x: x['date'], reverse=True)
        return anomalies[:10]
    elif widget_name == 'recent_readings':
        recent = []
        for meter in meters:
            last = meter.reading_set.order_by('-date').first()
            if last:
                recent.append({'meter': meter, 'date': last.date, 'consumption': float(last.total_consumption()), 'unit': meter.resource_type.unit})
        recent.sort(key=lambda x: x['date'], reverse=True)
        return recent[:10]
    return None


@login_required
def energy_dashboard(request):
    dashboard_settings = request.session.get('energy_dashboard_settings', {})
    selected_widgets = dashboard_settings.get('widgets', ['kpi', 'consumption_chart', 'resource_pie', 'top_meters', 'anomaly_table'])
    widget_order = dashboard_settings.get('order', [w for w in selected_widgets if w in AVAILABLE_WIDGETS])
    period = request.GET.get('period', dashboard_settings.get('period', 'month'))
    year_str = request.GET.get('year', dashboard_settings.get('year', date.today().year))
    month_str = request.GET.get('month', dashboard_settings.get('month', date.today().month))
    quarter_str = request.GET.get('quarter', dashboard_settings.get('quarter', (date.today().month - 1) // 3 + 1))
    resource_type = request.GET.get('resource_type', dashboard_settings.get('resource_type', 'all'))
    threshold_str = request.GET.get('threshold', dashboard_settings.get('threshold', 2.0))
    try:
        year = int(year_str) if year_str else date.today().year
    except ValueError:
        year = date.today().year
    try:
        month = int(month_str) if month_str else date.today().month
    except ValueError:
        month = date.today().month
    try:
        quarter = int(quarter_str) if quarter_str else (date.today().month - 1) // 3 + 1
    except ValueError:
        quarter = (date.today().month - 1) // 3 + 1
    try:
        threshold = float(threshold_str) if threshold_str else 2.0
    except ValueError:
        threshold = 2.0
    filters = {'period': period, 'year': year, 'month': month, 'quarter': quarter, 'resource_type': resource_type, 'threshold': threshold}
    widgets_data = {}
    for w in widget_order:
        if w in AVAILABLE_WIDGETS:
            widgets_data[w] = get_widget_data(w, request, filters)
    current_year = date.today().year
    years = range(current_year - 2, current_year + 1)
    months = [(1, 'Январь'), (2, 'Февраль'), (3, 'Март'), (4, 'Апрель'), (5, 'Май'), (6, 'Июнь'), (7, 'Июль'), (8, 'Август'), (9, 'Сентябрь'), (10, 'Октябрь'), (11, 'Ноябрь'), (12, 'Декабрь')]
    quarters = [1, 2, 3, 4]
    resource_types = ResourceType.objects.all()
    context = {
        'available_widgets': AVAILABLE_WIDGETS,
        'widgets_data': widgets_data,
        'widget_order': widget_order,
        'selected_widgets': selected_widgets,
        'filters': filters,
        'years': years,
        'months': months,
        'quarters': quarters,
        'resource_types': resource_types,
        'period_type': period,
        'selected_year': year,
        'selected_month': month,
        'selected_quarter': quarter,
        'selected_resource_type': resource_type,
        'anomaly_threshold': threshold,
    }
    return render(request, 'energy/dashboard.html', context)


@require_POST
@login_required
def save_dashboard_settings(request):
    data = json.loads(request.body)
    threshold = data.get('threshold', 2.0)
    if threshold == '':
        threshold = 2.0
    try:
        threshold = float(threshold)
    except (ValueError, TypeError):
        threshold = 2.0
    request.session['energy_dashboard_settings'] = {
        'widgets': data.get('widgets', []),
        'order': data.get('order', []),
        'period': data.get('period', 'month'),
        'year': int(data.get('year', date.today().year)) if data.get('year') else date.today().year,
        'month': int(data.get('month', date.today().month)) if data.get('month') else date.today().month,
        'quarter': int(data.get('quarter', (date.today().month - 1) // 3 + 1)) if data.get('quarter') else (date.today().month - 1) // 3 + 1,
        'resource_type': data.get('resource_type', 'all'),
        'threshold': threshold,
    }
    return JsonResponse({'status': 'ok'})


@login_required
def reset_initial_readings(request, pk):
    """
    Сброс начальных показаний счётчика (установка даты сброса и начальных значений для каждой зоны).
    """
    meter = get_object_or_404(Meter, pk=pk)
    if not can_edit_meter(request.user, meter):
        messages.error(request, "У вас нет прав на изменение начальных показаний.")
        return redirect('energy:meter_detail', pk=meter.pk)

    if request.method == 'POST':
        form = ResetInitialReadingsForm(meter, request.POST)
        if form.is_valid():
            reset_date = form.cleaned_data['reset_date']
            meter.reset_date = reset_date
            meter.save()

            if meter.is_multi_tariff:
                for key, value in form.cleaned_data.items():
                    if key.startswith('zone_'):
                        comp_id = int(key.split('_')[1])
                        comp = TariffComponent.objects.get(pk=comp_id)
                        # Удаляем все записи истории для этой зоны с date_from >= reset_date
                        InitialZoneValueHistory.objects.filter(
                            meter=meter,
                            tariff_component=comp,
                            date_from__gte=reset_date
                        ).delete()
                        # Создаём новую запись
                        InitialZoneValueHistory.objects.create(
                            meter=meter,
                            tariff_component=comp,
                            value=value,
                            date_from=reset_date
                        )
                meter.initial_value = 0
                meter.save()
            else:
                # Для однотарифного аналогично
                InitialValueHistory.objects.filter(
                    meter=meter,
                    date_from__gte=reset_date
                ).delete()
                InitialValueHistory.objects.create(
                    meter=meter,
                    value=form.cleaned_data['initial_value'],
                    date_from=reset_date
                )
                meter.initial_value = form.cleaned_data['initial_value']
                meter.save()

            meter.recalc_consumption()

            log_action(user=request.user, action='EDIT', model_name='Meter', object_id=meter.pk,
                       details=f"Сброс начальных показаний для счётчика {meter.serial_number} на дату {reset_date}", request=request)
            messages.success(request, "Начальные показания и дата сброса успешно обновлены.")
            return redirect('energy:meter_detail', pk=meter.pk)
    else:
        form = ResetInitialReadingsForm(meter)

    return render(request, 'energy/reset_initial_readings.html', {'form': form, 'meter': meter})