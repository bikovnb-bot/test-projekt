# buildings/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

from .models import Building, BuildingDocument, BuildingAppendix, BuildingOwnershipDocument
from .forms import (
    BuildingForm,
    BuildingRoomFormSet, BuildingSystemFormSet, BuildingLandscapingFormSet,
    BuildingInspectionFormSet, BuildingRepairFormSet, BuildingTenantFormSet,
    BuildingAppendixFormSet, BuildingDocumentFormSet, BuildingOwnershipDocumentFormSet
)
from users.models import UserRole


def can_edit_passport(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = getattr(user.profile, 'role', None) if hasattr(user, 'profile') else None
    return role in [UserRole.ADMIN, UserRole.ENGINEER, UserRole.DISPATCHER]


@login_required
def building_list(request):
    buildings = Building.objects.all().order_by('name', 'address')
    return render(request, 'buildings/building_list.html', {'buildings': buildings})


@login_required
def building_detail(request, pk):
    building = get_object_or_404(Building, pk=pk)
    return render(request, 'buildings/building_detail.html', {'building': building})


@login_required
def passport_detail(request, pk):
    building = get_object_or_404(Building, pk=pk)
    can_edit = can_edit_passport(request.user)
    context = {
        'building': building,
        'can_edit': can_edit,
        'rooms': building.rooms.all(),
        'systems': building.systems.all(),
        'landscaping': building.landscaping.all(),
        'inspections': building.inspections.all(),
        'repairs': building.repairs.all(),
        'tenants': building.tenants.all(),
        'appendixes': building.appendixes.all(),
        'documents': building.documents.all(),
        'ownership_docs': building.ownership_docs.all(),
        'sections': building.sections.all(),
    }
    return render(request, 'buildings/passport_detail.html', context)


@login_required
def passport_edit(request, pk):
    building = get_object_or_404(Building, pk=pk)
    if not can_edit_passport(request.user):
        messages.error(request, 'У вас нет прав на редактирование паспорта здания.')
        return redirect('buildings:passport_detail', pk=pk)

    if request.method == 'POST':
        form = BuildingForm(request.POST, instance=building)
        room_formset = BuildingRoomFormSet(request.POST, instance=building)
        system_formset = BuildingSystemFormSet(request.POST, instance=building)
        landscaping_formset = BuildingLandscapingFormSet(request.POST, instance=building)
        inspection_formset = BuildingInspectionFormSet(request.POST, instance=building)
        repair_formset = BuildingRepairFormSet(request.POST, instance=building)
        tenant_formset = BuildingTenantFormSet(request.POST, instance=building)
        appendix_formset = BuildingAppendixFormSet(request.POST, request.FILES, instance=building)
        document_formset = BuildingDocumentFormSet(request.POST, request.FILES, instance=building)
        ownership_formset = BuildingOwnershipDocumentFormSet(request.POST, request.FILES, instance=building)

        if (form.is_valid() and room_formset.is_valid() and system_formset.is_valid() and
            landscaping_formset.is_valid() and inspection_formset.is_valid() and
            repair_formset.is_valid() and tenant_formset.is_valid() and
            appendix_formset.is_valid() and document_formset.is_valid() and
            ownership_formset.is_valid()):
            with transaction.atomic():
                building = form.save()
                room_formset.save()
                system_formset.save()
                landscaping_formset.save()
                inspection_formset.save()
                repair_formset.save()
                tenant_formset.save()
                appendix_formset.save()
                document_formset.save()
                ownership_formset.save()
            messages.success(request, f'Паспорт здания "{building}" успешно обновлён.')
            return redirect('buildings:passport_detail', pk=building.pk)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = BuildingForm(instance=building)
        room_formset = BuildingRoomFormSet(instance=building)
        system_formset = BuildingSystemFormSet(instance=building)
        landscaping_formset = BuildingLandscapingFormSet(instance=building)
        inspection_formset = BuildingInspectionFormSet(instance=building)
        repair_formset = BuildingRepairFormSet(instance=building)
        tenant_formset = BuildingTenantFormSet(instance=building)
        appendix_formset = BuildingAppendixFormSet(instance=building)
        document_formset = BuildingDocumentFormSet(instance=building)
        ownership_formset = BuildingOwnershipDocumentFormSet(instance=building)

    context = {
        'form': form,
        'room_formset': room_formset,
        'system_formset': system_formset,
        'landscaping_formset': landscaping_formset,
        'inspection_formset': inspection_formset,
        'repair_formset': repair_formset,
        'tenant_formset': tenant_formset,
        'appendix_formset': appendix_formset,
        'document_formset': document_formset,
        'ownership_formset': ownership_formset,
        'building': building,
        'title': f'Редактирование паспорта: {building}',
    }
    return render(request, 'buildings/passport_edit.html', context)


@login_required
def passport_export_excel(request, pk):
    building = get_object_or_404(Building, pk=pk)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. Общие сведения
    ws_general = wb.create_sheet("Общие сведения")
    ws_general.append(['Параметр', 'Значение'])
    general_data = [
        ('Наименование учреждения', building.institution_name),
        ('Адрес', building.address),
        ('Кадастровый номер', building.cadastral_number),
        ('ФИО руководителя', building.director_name),
        ('Телефон', building.director_phone),
        ('Год постройки', building.year_built),
        ('Этажность', building.number_of_floors),
        ('Тип здания', building.get_building_type_display()),
        ('Количество помещений', building.number_of_rooms),
        ('Балансовая стоимость', building.balance_cost),
        ('Площадь территории', building.territory_area),
        ('Тип проекта', building.project_type),
    ]
    for row in general_data:
        ws_general.append(row)
    ws_general.column_dimensions['A'].width = 30
    ws_general.column_dimensions['B'].width = 50

    # 2. Конструктивная характеристика
    ws_construct = wb.create_sheet("Конструктивная характеристика")
    ws_construct.append(['Элемент', 'Описание'])
    construct_data = [
        ('Фундаменты', building.foundation_desc),
        ('Несущий каркас', building.frame_desc),
        ('Стены и перегородки', building.walls_desc),
        ('Перекрытия', building.floors_desc),
        ('Лестницы', building.stairs_desc),
        ('Несущие элементы кровли', building.roof_structure_desc),
        ('Кровля', building.roof_cover_desc),
    ]
    for row in construct_data:
        ws_construct.append(row)
    ws_construct.column_dimensions['A'].width = 30
    ws_construct.column_dimensions['B'].width = 50

    # 3. Площади (фиксированная таблица)
    ws_areas = wb.create_sheet("Площади")
    ws_areas.append(['Тип помещения', 'Общая площадь (м²)', 'Жилая площадь (м²)'])
    ws_areas.append(['Общая площадь здания', building.total_area, '—'])
    ws_areas.append(['Жилые помещения', building.residential_area, building.residential_livable_area])
    ws_areas.append(['Нежилые помещения', building.non_residential_area, '—'])
    for col in ws_areas.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_areas.column_dimensions[col_letter].width = 25

    # 4. Внутренние помещения (с группировкой по секциям)
    ws_rooms = wb.create_sheet("Внутренние помещения")
    ws_rooms.append(['Часть здания', 'Этаж', 'Наименование', 'Потолок', 'Стены', 'Полы', 'Окна', 'Двери', 'Вид ремонта', 'Год ремонта', 'Состояние', 'Рекомендации'])
    for room in building.rooms.all():
        ws_rooms.append([
            room.section.name if room.section else '',
            room.get_floor_display(),
            room.name,
            room.ceiling_finish,
            room.walls_finish,
            room.floors_finish,
            room.windows,
            room.doors,
            room.last_repair_type,
            room.last_repair_year,
            room.condition,
            room.recommendations,
        ])
    for col in ws_rooms.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_rooms.column_dimensions[col_letter].width = 20

    # 5. Инженерные системы
    ws_systems = wb.create_sheet("Инженерные системы")
    sys_headers = ['Часть здания', 'Тип системы', 'Описание', 'Мощность', 'Диаметр ввода', '№ района', 'Телефон', '№ абонента',
                   'Счётчик1 (тип/№/поверка)', 'Счётчик2', 'Счётчик3', 'Характеристика сети', 'Последний ремонт']
    ws_systems.append(sys_headers)
    for sys in building.systems.all():
        meter1 = f"{sys.meter1_type} / {sys.meter1_number} / {sys.meter1_verification_date}" if sys.meter1_type else ''
        meter2 = f"{sys.meter2_type} / {sys.meter2_number} / {sys.meter2_verification_date}" if sys.meter2_type else ''
        meter3 = f"{sys.meter3_type} / {sys.meter3_number} / {sys.meter3_verification_date}" if sys.meter3_type else ''
        repair = f"{sys.last_repair_type} ({sys.last_repair_year})" if sys.last_repair_type else ''
        ws_systems.append([
            sys.section.name if sys.section else '',
            sys.get_system_type_display(),
            sys.type_desc,
            sys.power,
            sys.inlet_diameter,
            sys.district_number,
            sys.district_phone,
            sys.subscriber_number,
            meter1, meter2, meter3,
            sys.network_description,
            repair
        ])
    for col in ws_systems.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_systems.column_dimensions[col_letter].width = 20

    # 6. Благоустройство
    ws_land = wb.create_sheet("Благоустройство")
    ws_land.append(['Часть здания', 'Элемент', 'Кол-во/площадь', 'Характеристика', 'Наличие документов', 'Состояние', 'Рекомендации'])
    for l in building.landscaping.all():
        ws_land.append([
            l.section.name if l.section else '',
            l.element,
            l.quantity,
            l.characteristic,
            'Да' if l.has_documents else 'Нет',
            l.condition,
            l.recommendations
        ])
    for col in ws_land.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_land.column_dimensions[col_letter].width = 20

    # 7. Проверки
    ws_inspections = wb.create_sheet("Проверки")
    ws_inspections.append(['Дата', 'Часть здания', 'Причина', 'Проверяющий', 'Замечания', 'Заключение', 'Рекомендации'])
    for i in building.inspections.all():
        ws_inspections.append([
            i.inspection_date,
            i.section.name if i.section else '',
            i.reason,
            i.inspector,
            i.findings,
            i.conclusion,
            i.recommendations
        ])
    for col in ws_inspections.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_inspections.column_dimensions[col_letter].width = 20

    # 8. Ремонты
    ws_repairs = wb.create_sheet("Ремонты")
    ws_repairs.append(['Часть здания', 'Объект', 'Вид', 'Начало', 'Окончание', '№ контракта', 'Сумма', 'Подрядчик', 'Гарантия'])
    for r in building.repairs.all():
        ws_repairs.append([
            r.section.name if r.section else '',
            r.object_name,
            r.repair_type,
            r.start_date,
            r.end_date,
            r.contract_number,
            r.contract_amount,
            r.contractor,
            r.warranty_period
        ])
    for col in ws_repairs.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_repairs.column_dimensions[col_letter].width = 20

    # 9. Арендаторы
    ws_tenants = wb.create_sheet("Арендаторы")
    ws_tenants.append(['Часть здания', 'Наименование', 'Вид деятельности', 'Арендуемая площадь', 'Срок аренды', 'Договор'])
    for t in building.tenants.all():
        ws_tenants.append([
            t.section.name if t.section else '',
            t.name,
            t.activity,
            t.rented_areas,
            t.lease_term,
            t.contract_number
        ])
    for col in ws_tenants.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_tenants.column_dimensions[col_letter].width = 20

    # 10. Приложения (список файлов)
    ws_apps = wb.create_sheet("Приложения")
    ws_apps.append(['Часть здания', 'Тип', 'Название', 'Файл'])
    for a in building.appendixes.all():
        ws_apps.append([
            a.section.name if a.section else '',
            a.get_appendix_type_display(),
            a.title or '',
            a.file.name if a.file else ''
        ])
    for col in ws_apps.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_apps.column_dimensions[col_letter].width = 30

    # 11. Документы
    ws_docs = wb.create_sheet("Документы")
    ws_docs.append(['Часть здания', 'Название', 'Файл'])
    for d in building.documents.all():
        ws_docs.append([
            d.section.name if d.section else '',
            d.title or '',
            d.file.name if d.file else ''
        ])
    for col in ws_docs.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_docs.column_dimensions[col_letter].width = 30

    # 12. Документы о собственности
    ws_ownership = wb.create_sheet("Документы о собственности")
    ws_ownership.append(['Название', 'Файл', 'Дата загрузки'])
    for o in building.ownership_docs.all():
        ws_ownership.append([
            o.title or '',
            o.file.name if o.file else '',
            o.uploaded_at.strftime('%d.%m.%Y %H:%M') if o.uploaded_at else ''
        ])
    for col in ws_ownership.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_ownership.column_dimensions[col_letter].width = 30

    # Автоширина для всех листов
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len+2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="passport_{building.id}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@require_POST
@login_required
def delete_building_file(request, pk, file_type, file_id):
    building = get_object_or_404(Building, pk=pk)
    if not can_edit_passport(request.user):
        return JsonResponse({'error': 'Нет прав'}, status=403)
    if file_type == 'document':
        file_obj = get_object_or_404(BuildingDocument, pk=file_id, building=building)
    elif file_type == 'appendix':
        file_obj = get_object_or_404(BuildingAppendix, pk=file_id, building=building)
    elif file_type == 'ownership':
        file_obj = get_object_or_404(BuildingOwnershipDocument, pk=file_id, building=building)
    else:
        return JsonResponse({'error': 'Неверный тип'}, status=400)
    file_obj.file.delete()
    file_obj.delete()
    return JsonResponse({'success': True})