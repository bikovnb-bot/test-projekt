import pandas as pd
import re
from openpyxl import load_workbook

# Загружаем файл
file_path = 'Остатки ОС Быков Н.Б..xlsx'
wb = load_workbook(file_path, data_only=True)
sheet = wb.active

# Список для сбора всех записей
records = []

# Переменные для хранения текущего контекста
current_account = None
current_kfo = None
current_kps = None
current_tmo_employee = None
current_tmo_location = None
current_bal_cost = None
current_quantity = None
current_amort = None
current_impair = None
current_residual = None
current_initial = None

# Регулярка для определения строки счёта (начинается с цифр и точки)
account_pattern = re.compile(r'^\d+\.\d+,')

# Читаем строки начиная с первой (индекс 0)
for row in sheet.iter_rows(min_row=1, values_only=True):
    # если строка пустая – пропускаем
    if all(cell is None for cell in row):
        continue

    # берём первый столбец (А) как основной индикатор
    first_cell = row[0] if row[0] is not None else ''
    first_cell = str(first_cell).strip()

    # 1. Проверяем, не является ли строка заголовком счёта
    if account_pattern.match(first_cell):
        # Сохраняем предыдущий объект, если он был (обычно это строки с группировкой)
        # Но сами данные объектов идут позже
        current_account = first_cell
        # В этой же строке могут быть итоговые значения по счёту:
        # Балансовая стоимость – столбец T (индекс 19), Количество – U (20), Амортизация – V (21), Обесценение – W (22), Остаточная – X (23), Первоначальная – Y (24)
        # Индексы считаем с 0: A=0, B=1, ... T=19, U=20, V=21, W=22, X=23, Y=24
        try:
            current_bal_cost = row[19] if len(row) > 19 else None
            current_quantity = row[20] if len(row) > 20 else None
            current_amort = row[21] if len(row) > 21 else None
            current_impair = row[22] if len(row) > 22 else None
            current_residual = row[23] if len(row) > 23 else None
            current_initial = row[24] if len(row) > 24 else None
        except:
            pass
        continue

    # 2. Проверяем, не является ли строка подзаголовком КФО, КПС, ЦМО
    # Они обычно имеют значения в первых столбцах, например "2", "07060000000000000", "Быков Николай Борисович", "Инженерно-техническая служба"
    # Но они могут повторяться внутри одного счета. Будем их считывать последовательно.
    # Если строка содержит только одно значение в первом столбце и остальные пустые – это скорее всего подзаголовок.
    # Но у нас есть строки, где после счета идут именно такие подзаголовки.
    # Лучше определять по номеру столбца: если в столбце B (индекс 1) пусто, а в A что-то есть – возможно это КФО/КПС/ЦМО.
    # Однако есть строки с номером п/п – они имеют значение в A (цифра) и заполнены остальные столбцы.
    # Поэтому проверим: если в B пусто, но в A есть непустое значение и оно не является числом (номер п/п) – это подзаголовок.
    # Но номера п/п – это числа, а КФО может быть "2" – тоже число. Значит, нужно смотреть на наличие данных в других столбцах.
    # Проще: если в столбце C (индекс 2) пусто, а в A непусто – это скорее подзаголовок (так как у объектов название в C).
    # Проверим.
    if row[1] is None and row[2] is None:
        # Возможно это подзаголовок
        # Определим, что это: КФО, КПС, ФИО, Место хранения – по порядку они идут.
        # Мы можем запомнить последние установленные значения.
        if first_cell.isdigit() and len(first_cell) <= 2:
            # скорее всего КФО
            current_kfo = first_cell
        elif len(first_cell) > 10 and first_cell.isdigit():
            # КПС – длинный цифровой код
            current_kps = first_cell
        elif 'Быков' in first_cell or 'Николай' in first_cell:
            current_tmo_employee = first_cell
        elif 'служба' in first_cell or 'техническая' in first_cell:
            current_tmo_location = first_cell
        else:
            # возможно что-то ещё, но пока пропускаем
            pass
        continue

    # 3. Теперь строки с объектами: они имеют номер п/п в A (цифра) и название в C (индекс 2)
    if row[0] is not None and str(row[0]).strip().isdigit():
        # Извлекаем данные
        item = {
            'счет': current_account,
            'кфо': current_kfo,
            'кпс': current_kps,
            'цмо_сотрудник': current_tmo_employee,
            'цмо_место': current_tmo_location,
            'номер_пп': row[0],
            'наименование': row[2] if len(row) > 2 else None,
            'инвентарный_номер': row[8] if len(row) > 8 else None,
            'ок_of': row[9] if len(row) > 9 else None,
            'амортизационная_группа': row[11] if len(row) > 11 else None,
            'способ_начисления': row[12] if len(row) > 12 else None,
            'дата_принятия': row[13] if len(row) > 13 else None,
            'состояние': row[14] if len(row) > 14 else None,
            'срок_полезного_использования': row[15] if len(row) > 15 else None,
            'мес_норма_износа': row[16] if len(row) > 16 else None,
            'износ_процент': row[17] if len(row) > 17 else None,
            'балансовая_стоимость': row[19] if len(row) > 19 else current_bal_cost,
            'количество': row[20] if len(row) > 20 else current_quantity,
            'сумма_амортизации': row[21] if len(row) > 21 else current_amort,
            'сумма_обесценения': row[22] if len(row) > 22 else current_impair,
            'остаточная_стоимость': row[23] if len(row) > 23 else current_residual,
            'первоначальная_стоимость': row[24] if len(row) > 24 else current_initial,
        }
        records.append(item)

# Преобразуем в DataFrame
df = pd.DataFrame(records)

# Сохраняем в CSV или JSON
df.to_excel('imported_os.xlsx', index=False, engine='openpyxl')

print(f'Импортировано {len(df)} объектов')
print(df.head())