# assets/management/commands/import_assets.py
import re
from datetime import datetime
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from openpyxl import load_workbook

from assets.models import Asset, AssetCategory


class Command(BaseCommand):
    help = 'Импорт имущества из Excel-файла (imported_os.xlsx) с автоматическим заполнением пропусков'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к Excel-файлу')

    def handle(self, *args, **options):
        file_path = options['file_path']
        self.stdout.write(f'Начинаем импорт из файла: {file_path}')
        
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            self.stderr.write(f'Ошибка открытия файла: {e}')
            return

        category_cache = {}
        responsible_user = self._get_responsible_user()
        
        created_count = 0
        errors = []
        auto_fields = {
            'inventory_number': 0,
            'name': 0,
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            account = row[0] if len(row) > 0 and row[0] else ''
            employee = row[1] if len(row) > 1 and row[1] else ''
            name = row[2] if len(row) > 2 and row[2] else ''
            inventory_number = row[3] if len(row) > 3 and row[3] else ''
            date_str = row[4] if len(row) > 4 and row[4] else ''
            useful_life_str = row[5] if len(row) > 5 and row[5] else ''
            cost_str = row[6] if len(row) > 6 and row[6] else ''
            quantity = row[7] if len(row) > 7 and row[7] else 1

            # Инвентарный номер – если он пустой, модель сама сгенерирует
            inv_num_str = str(inventory_number).strip() if inventory_number else ''
            if not inv_num_str:
                # Оставляем пустую строку – модель сгенерирует номер автоматически
                auto_fields['inventory_number'] += 1

            # Наименование
            name_str = str(name).strip() if name else ''
            if not name_str:
                name_str = f"Без названия (инв. будет сгенерирован)"
                auto_fields['name'] += 1

            # Парсинг даты
            purchase_date = self._parse_date(date_str)

            # Парсинг срока полезного использования
            useful_life_months = self._parse_useful_life(useful_life_str)

            # Парсинг стоимости
            cost = self._parse_cost(cost_str)

            # Извлечение числового кода категории
            category_code = self._extract_category_code(account)
            if not category_code:
                category_code = 'Без категории'
            category = self._get_category(category_code, account, category_cache)

            # Создаём запись
            try:
                asset = Asset(
                    inventory_number=inv_num_str,  # может быть пустым
                    name=name_str,
                    category=category,
                    description='Импортировано из Excel',
                    purchase_date=purchase_date,
                    useful_life_months=useful_life_months,
                    cost=cost,
                    responsible_person=responsible_user,
                    status='in_use',
                    notes=f'Количество: {quantity}; Источник: {employee}',
                    imported_from_excel=True,
                )
                asset.save()
                created_count += 1
                if created_count % 50 == 0:
                    self.stdout.write(f'Создано записей: {created_count}')
            except Exception as e:
                errors.append(f'Строка {row_idx}: ошибка сохранения: {e}')
                continue

        self.stdout.write(self.style.SUCCESS(
            f'Импорт завершён. Создано: {created_count}.'
        ))
        self.stdout.write(f'Автоматически заполнено: инвентарных номеров: {auto_fields["inventory_number"]}, названий: {auto_fields["name"]}.')
        if errors:
            self.stdout.write(self.style.WARNING(f'Ошибки ({len(errors)}): {", ".join(errors[:10])}'))

    def _get_responsible_user(self):
        user = User.objects.filter(first_name='Николай', last_name='Быков').first()
        if not user:
            self.stdout.write('Пользователь "Николай Быков" не найден. Создаём...')
            user = User.objects.create_user(
                username='bikov',
                first_name='Николай',
                last_name='Быков',
                password='password123'
            )
            self.stdout.write(f'Создан пользователь: {user.get_full_name()} (логин: bikov)')
        else:
            self.stdout.write(f'Найден пользователь: {user.get_full_name()}')
        return user

    def _parse_date(self, date_str):
        if not date_str:
            return None
        try:
            if isinstance(date_str, datetime):
                return date_str.date()
            if isinstance(date_str, str):
                date_str_clean = date_str.strip()
                if re.match(r'\d{2}\.\d{2}\.\d{4}', date_str_clean):
                    return datetime.strptime(date_str_clean, '%d.%m.%Y').date()
                elif re.match(r'\d{4}-\d{2}-\d{2}', date_str_clean):
                    return datetime.strptime(date_str_clean, '%Y-%m-%d').date()
                elif re.match(r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', date_str_clean):
                    return datetime.strptime(date_str_clean, '%Y-%m-%d %H:%M:%S').date()
        except Exception as e:
            self.stdout.write(f'Ошибка парсинга даты "{date_str}": {e}')
        return None

    def _parse_useful_life(self, value):
        if not value:
            return None
        try:
            return int(float(str(value).replace(',', '.')))
        except:
            return None

    def _parse_cost(self, value):
        if not value:
            return None
        try:
            return Decimal(str(value).replace(',', '.'))
        except:
            return None

    def _extract_category_code(self, account):
        if not account:
            return None
        account_str = str(account).strip()
        match = re.search(r'^(\d+\.\d+)', account_str)
        if match:
            return match.group(1)
        return None

    def _get_category(self, category_code, original_account, cache):
        if category_code not in cache:
            category, created = AssetCategory.objects.get_or_create(
                name=category_code,
                defaults={'description': f'Импортировано из Excel (оригинал: {original_account})'}
            )
            cache[category_code] = category
            if created:
                self.stdout.write(f'Создана категория: {category.name}')
        return cache[category_code]