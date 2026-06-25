# assets/views.py
# Полная версия с поддержкой импорта/экспорта, офлайн-инвентаризации, API, фото, галереи, пагинации и сортировки

import logging
import re
from decimal import Decimal
from datetime import datetime

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count
from django.utils import timezone

# REST Framework для API
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from users.models import UserRole
from .models import Asset, AssetCategory, AssetAssignment, AssetCheck, AssetPhoto
from .forms import AssetForm, AssetAssignmentForm, AssetCheckForm, AssetGalleryForm
from .serializers import AssetSimpleSerializer, InventorySyncSerializer

logger = logging.getLogger(__name__)


# ---------- PERMISSION HELPERS ----------
def can_edit_asset(user, asset=None):
    if user.is_superuser:
        return True
    role = getattr(user, 'profile', None)
    if role and role.role in [UserRole.ADMIN, UserRole.ENGINEER]:
        return True
    return False

def can_delete_asset(user):
    if user.is_superuser:
        return True
    role = getattr(user, 'profile', None)
    return role and role.role == UserRole.ADMIN

def can_assign_asset(user):
    if user.is_superuser:
        return True
    role = getattr(user, 'profile', None)
    return role and role.role in [UserRole.ADMIN, UserRole.ENGINEER]


# ---------- CRUD (CBV) ----------
class AssetListView(LoginRequiredMixin, ListView):
    model = Asset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 20

    def get_queryset(self):
        qs = Asset.objects.select_related('category', 'responsible_person').all()
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(
                Q(inventory_number__icontains=search) |
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(location__icontains=search)
            )
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        category_id = self.request.GET.get('category_id')
        if category_id and category_id != 'all':
            qs = qs.filter(category_id=category_id)
        responsible = self.request.GET.get('responsible')
        if responsible:
            qs = qs.filter(responsible_person_id=responsible)

        # Сортировка
        sort_field = self.request.GET.get('sort', 'created_at')
        order = self.request.GET.get('order', 'desc')
        if order == 'desc':
            sort_field = f'-{sort_field}'
        qs = qs.order_by(sort_field)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        categories = AssetCategory.objects.annotate(
            asset_count=Count('assets')
        ).order_by('name')
        context['categories'] = categories
        context['selected_category_id'] = self.request.GET.get('category_id', 'all')
        base_query = self.request.GET.copy()
        base_query.pop('category_id', None)
        context['base_query'] = base_query.urlencode()
        context['total_assets'] = Asset.objects.count()
        context['status_choices'] = Asset.STATUS_CHOICES
        context['users'] = User.objects.filter(is_active=True).order_by('username')
        context['selected_status'] = self.request.GET.get('status')
        context['selected_responsible'] = self.request.GET.get('responsible')
        context['search'] = self.request.GET.get('search')
        context['can_edit'] = can_edit_asset(self.request.user)
        context['can_delete'] = can_delete_asset(self.request.user)
        context['can_assign'] = can_assign_asset(self.request.user)

        # Параметры сортировки
        context['current_sort'] = self.request.GET.get('sort', 'created_at')
        context['current_order'] = self.request.GET.get('order', 'desc')
        sort_params = self.request.GET.copy()
        context['sort_params'] = sort_params
        return context


class AssetDetailView(LoginRequiredMixin, DetailView):
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.get_object()
        context['assignments'] = asset.assignments.all().order_by('-assigned_at')
        context['checks'] = asset.checks.all().order_by('-checked_at')
        context['photos'] = asset.photos.all().order_by('order', 'uploaded_at')
        context['can_edit'] = can_edit_asset(self.request.user, asset)
        context['can_delete'] = can_delete_asset(self.request.user)
        context['can_assign'] = can_assign_asset(self.request.user)
        context['gallery_form'] = AssetGalleryForm()
        return context


class AssetCreateView(LoginRequiredMixin, CreateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('assets:asset_list')

    def dispatch(self, request, *args, **kwargs):
        if not can_edit_asset(request.user):
            messages.error(request, 'У вас нет прав на создание имущества.')
            return redirect('assets:asset_list')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Имущество успешно создано.')
        return super().form_valid(form)


class AssetUpdateView(LoginRequiredMixin, UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('assets:asset_list')

    def dispatch(self, request, *args, **kwargs):
        if not can_edit_asset(request.user, self.get_object()):
            messages.error(request, 'У вас нет прав на редактирование этого имущества.')
            return redirect('assets:asset_detail', pk=self.get_object().pk)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Имущество успешно обновлено.')
        return super().form_valid(form)


class AssetDeleteView(LoginRequiredMixin, DeleteView):
    model = Asset
    template_name = 'assets/asset_confirm_delete.html'
    success_url = reverse_lazy('assets:asset_list')

    def dispatch(self, request, *args, **kwargs):
        if not can_delete_asset(request.user):
            messages.error(request, 'У вас нет прав на удаление имущества.')
            return redirect('assets:asset_list')
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Имущество успешно удалено.')
        return super().delete(request, *args, **kwargs)


# ---------- ДЕЙСТВИЯ (FBV) ----------
@login_required
def assign_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if not can_assign_asset(request.user):
        messages.error(request, 'У вас нет прав на закрепление имущества.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if request.method == 'POST':
        form = AssetAssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.asset = asset
            asset.responsible_person = assignment.assigned_to
            asset.save()
            assignment.save()
            messages.success(request, f'Имущество закреплено за {assignment.assigned_to}.')
            return redirect('assets:asset_detail', pk=asset.pk)
    else:
        form = AssetAssignmentForm(initial={'assigned_to': asset.responsible_person})
    return render(request, 'assets/assign_asset.html', {'form': form, 'asset': asset})


@login_required
def return_asset(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if not can_assign_asset(request.user):
        messages.error(request, 'У вас нет прав на возврат имущества.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if request.method == 'POST':
        assignment = asset.assignments.filter(returned_at__isnull=True).first()
        if assignment:
            assignment.returned_at = timezone.now()
            assignment.save()
        asset.responsible_person = None
        asset.save()
        messages.success(request, 'Имущество возвращено на склад.')
        return redirect('assets:asset_detail', pk=asset.pk)
    return render(request, 'assets/return_asset.html', {'asset': asset})


@login_required
def add_asset_check(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if not can_edit_asset(request.user, asset):
        messages.error(request, 'У вас нет прав на проведение проверки.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if request.method == 'POST':
        form = AssetCheckForm(request.POST)
        if form.is_valid():
            check = form.save(commit=False)
            check.asset = asset
            check.checked_by = request.user
            check.save()
            messages.success(request, 'Проверка сохранена.')
            return redirect('assets:asset_detail', pk=asset.pk)
    else:
        form = AssetCheckForm()
    return render(request, 'assets/add_check.html', {'form': form, 'asset': asset})


@login_required
def generate_asset_qr(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if not asset.qr_code:
        asset.generate_qr_code()
        asset.save()
    return redirect('assets:asset_detail', pk=asset.pk)


@login_required
def download_asset_qr(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if not asset.qr_code:
        messages.error(request, 'QR-код не сгенерирован.')
        return redirect('assets:asset_detail', pk=asset.pk)
    response = HttpResponse(asset.qr_code, content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{asset.inventory_number}.png"'
    return response


@login_required
def asset_report(request):
    if not can_edit_asset(request.user):
        messages.error(request, 'У вас нет доступа к отчётам.')
        return redirect('assets:asset_list')
    qs = Asset.objects.select_related('category', 'responsible_person').all()
    context = {
        'assets': qs,
        'total': qs.count(),
        'by_status': {s: qs.filter(status=s).count() for s, _ in Asset.STATUS_CHOICES},
    }
    return render(request, 'assets/asset_report.html', context)


# ---------- ИНВЕНТАРИЗАЦИЯ ----------
@login_required
def inventory_asset(request, pk):
    """Отметка имущества в инвентаризации (вызывается по QR-ссылке)."""
    asset = get_object_or_404(Asset, pk=pk)
    today = timezone.now().date()
    existing = AssetCheck.objects.filter(
        asset=asset,
        checked_by=request.user,
        checked_at__date=today
    ).first()
    if existing:
        messages.info(request, f'Имущество "{asset.name}" уже было отмечено сегодня.')
    else:
        check = AssetCheck(
            asset=asset,
            checked_by=request.user,
            condition='good',
            notes=f"Отмечено при инвентаризации {timezone.now().strftime('%d.%m.%Y %H:%M')}"
        )
        check.save()
        messages.success(request, f'Имущество "{asset.name}" отмечено в инвентаризации.')
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('assets:asset_detail', pk=asset.pk)


@login_required
def inventory_scan(request):
    """Страница инвентаризации с камерой и списком отсканированных объектов."""
    today = timezone.now().date()
    scans_today = AssetCheck.objects.filter(
        checked_by=request.user,
        checked_at__date=today
    ).select_related('asset').order_by('-checked_at')
    return render(request, 'assets/inventory_scan.html', {'scans': scans_today})


@login_required
def inventory_scan_ajax(request):
    """
    AJAX-обработчик сканирования QR-кода.
    Принимает POST с inventory_number, создаёт запись проверки и возвращает данные актива.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)

    inventory_number = request.POST.get('inventory_number', '').strip()
    if not inventory_number:
        return JsonResponse({'success': False, 'error': 'Не указан инвентарный номер'})

    asset = Asset.objects.filter(inventory_number=inventory_number).first()
    if not asset:
        return JsonResponse({'success': False, 'error': 'Имущество не найдено'})

    today = timezone.now().date()
    existing = AssetCheck.objects.filter(
        asset=asset,
        checked_by=request.user,
        checked_at__date=today
    ).first()
    if existing:
        return JsonResponse({
            'success': False,
            'error': 'Уже отмечено сегодня',
            'asset': {
                'id': asset.id,
                'name': asset.name,
                'inventory_number': asset.inventory_number,
                'already_scanned': True
            }
        })

    check = AssetCheck(
        asset=asset,
        checked_by=request.user,
        condition='good',
        notes=f"Отмечено при инвентаризации {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    )
    check.save()

    return JsonResponse({
        'success': True,
        'asset': {
            'id': asset.id,
            'name': asset.name,
            'inventory_number': asset.inventory_number,
            'scanned_at': check.checked_at.strftime('%d.%m.%Y %H:%M'),
            'checked_by': request.user.get_full_name() or request.user.username,
        }
    })


@login_required
def inventory_history(request):
    """История инвентаризаций (список всех проверок)."""
    checks = AssetCheck.objects.select_related('asset', 'checked_by').order_by('-checked_at')
    paginator = Paginator(checks, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'assets/inventory_history.html', {'page_obj': page_obj})


# ---------- API ДЛЯ ОФЛАЙН-ИНВЕНТАРИЗАЦИИ ----------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def asset_list_api(request):
    """Возвращает список всех активов (id, инвентарный номер, название)."""
    assets = Asset.objects.all().only('id', 'inventory_number', 'name')
    serializer = AssetSimpleSerializer(assets, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def inventory_sync_api(request):
    """Принимает список инвентарных номеров и создаёт записи AssetCheck."""
    serializer = InventorySyncSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    inventory_numbers = serializer.validated_data['inventory_numbers']
    user = request.user
    created_count = 0
    errors = []

    for inv_number in inventory_numbers:
        try:
            asset = Asset.objects.get(inventory_number=inv_number)
        except Asset.DoesNotExist:
            errors.append(f'Не найден актив с номером {inv_number}')
            continue

        today = timezone.now().date()
        if AssetCheck.objects.filter(asset=asset, checked_by=user, checked_at__date=today).exists():
            continue

        AssetCheck.objects.create(
            asset=asset,
            checked_by=user,
            condition='good',
            notes=f'Отмечено офлайн {timezone.now().strftime("%d.%m.%Y %H:%M")}'
        )
        created_count += 1

    return Response({
        'success': True,
        'created': created_count,
        'errors': errors
    })


# ---------- ИМПОРТ / ЭКСПОРТ ----------
@login_required
def export_assets(request):
    """Экспорт всех объектов Asset в Excel"""
    if not can_edit_asset(request.user):
        messages.error(request, 'У вас нет прав на экспорт.')
        return redirect('assets:asset_list')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Имущество'

    headers = [
        'Счет (категория)', 'Ответственный (ФИО)', 'Наименование', 'Инвентарный номер',
        'Дата ввода (дд.мм.гггг)', 'Срок полезного использования (мес.)',
        'Балансовая стоимость (руб.)', 'Количество', 'Серийный номер', 'Модель',
        'Производитель', 'Местонахождение', 'Статус', 'Примечания'
    ]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    assets = Asset.objects.select_related('category', 'responsible_person').all()
    for asset in assets:
        status_display = dict(Asset.STATUS_CHOICES).get(asset.status, asset.status)
        row = [
            asset.category.name if asset.category else '',
            asset.responsible_person.get_full_name() if asset.responsible_person else '',
            asset.name,
            asset.inventory_number,
            asset.purchase_date.strftime('%d.%m.%Y') if asset.purchase_date else '',
            asset.useful_life_months or '',
            str(asset.cost) if asset.cost else '',
            1,
            asset.serial_number or '',
            asset.model or '',
            asset.manufacturer or '',
            asset.location or '',
            status_display,
            asset.notes or '',
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="assets_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


@login_required
def import_assets_view(request):
    """Импорт имущества из Excel с обновлением существующих записей и детальным отчётом об ошибках"""
    if not can_edit_asset(request.user):
        messages.error(request, 'У вас нет прав на импорт.')
        return redirect('assets:asset_list')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        file = request.FILES['excel_file']
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f'Не удалось открыть файл: {e}')
            return redirect('assets:import_assets')

        headers = [cell.value for cell in ws[1] if cell.value]
        required_headers = ['Счет (категория)', 'Наименование', 'Инвентарный номер']
        for req in required_headers:
            if req not in headers:
                messages.error(request, f'В файле отсутствует обязательная колонка: "{req}"')
                return redirect('assets:import_assets')

        col_map = {
            'category': headers.index('Счет (категория)') if 'Счет (категория)' in headers else -1,
            'responsible': headers.index('Ответственный (ФИО)') if 'Ответственный (ФИО)' in headers else -1,
            'name': headers.index('Наименование'),
            'inventory': headers.index('Инвентарный номер'),
            'purchase_date': headers.index('Дата ввода (дд.мм.гггг)') if 'Дата ввода (дд.мм.гггг)' in headers else -1,
            'useful_life': headers.index('Срок полезного использования (мес.)') if 'Срок полезного использования (мес.)' in headers else -1,
            'cost': headers.index('Балансовая стоимость (руб.)') if 'Балансовая стоимость (руб.)' in headers else -1,
            'quantity': headers.index('Количество') if 'Количество' in headers else -1,
            'serial': headers.index('Серийный номер') if 'Серийный номер' in headers else -1,
            'model': headers.index('Модель') if 'Модель' in headers else -1,
            'manufacturer': headers.index('Производитель') if 'Производитель' in headers else -1,
            'location': headers.index('Местонахождение') if 'Местонахождение' in headers else -1,
            'status': headers.index('Статус') if 'Статус' in headers else -1,
            'notes': headers.index('Примечания') if 'Примечания' in headers else -1,
        }

        created_count = 0
        updated_count = 0
        errors = []
        warnings = []
        category_cache = {}
        user_cache = {}
        seen_inventories = set()  # для проверки дубликатов внутри файла

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            inventory_number = str(row[col_map['inventory']]).strip() if col_map['inventory'] != -1 and row[col_map['inventory']] else ''
            if not inventory_number:
                errors.append(f'Строка {row_idx}: пропущен инвентарный номер')
                continue

            # Проверка уникальности: дубли в файле
            if inventory_number in seen_inventories:
                errors.append(f'Строка {row_idx}: дубликат инвентарного номера "{inventory_number}" в файле')
                continue
            seen_inventories.add(inventory_number)

            # Проверка существования в БД (уникальность)
            if Asset.objects.filter(inventory_number=inventory_number).exists():
                errors.append(f'Строка {row_idx}: инвентарный номер "{inventory_number}" уже существует в базе')
                continue

            # --- Обработка полей ---
            category_name = str(row[col_map['category']]).strip() if col_map['category'] != -1 and row[col_map['category']] else ''
            category = None
            if category_name:
                if category_name not in category_cache:
                    cat, _ = AssetCategory.objects.get_or_create(name=category_name)
                    category_cache[category_name] = cat
                category = category_cache[category_name]

            responsible = None
            if col_map['responsible'] != -1 and row[col_map['responsible']]:
                full_name = str(row[col_map['responsible']]).strip()
                if full_name:
                    if full_name not in user_cache:
                        name_parts = full_name.split(' ', 1)
                        if len(name_parts) == 2:
                            user = User.objects.filter(first_name=name_parts[0], last_name=name_parts[1]).first()
                        else:
                            user = User.objects.filter(username=full_name).first()
                        user_cache[full_name] = user
                    responsible = user_cache[full_name]
                    if not responsible:
                        warnings.append(f'Строка {row_idx}: пользователь "{full_name}" не найден, поле ответственного оставлено пустым')

            name = str(row[col_map['name']]).strip() if col_map['name'] != -1 and row[col_map['name']] else ''
            if not name:
                errors.append(f'Строка {row_idx}: отсутствует наименование')
                continue

            purchase_date = None
            if col_map['purchase_date'] != -1 and row[col_map['purchase_date']]:
                date_str = str(row[col_map['purchase_date']]).strip()
                try:
                    if re.match(r'\d{2}\.\d{2}\.\d{4}', date_str):
                        purchase_date = datetime.strptime(date_str, '%d.%m.%Y').date()
                    else:
                        warnings.append(f'Строка {row_idx}: неверный формат даты "{date_str}" (ожидается дд.мм.гггг)')
                except Exception:
                    warnings.append(f'Строка {row_idx}: не удалось распарсить дату "{date_str}"')

            useful_life_months = None
            if col_map['useful_life'] != -1 and row[col_map['useful_life']]:
                try:
                    useful_life_months = int(float(str(row[col_map['useful_life']]).replace(',', '.')))
                except:
                    warnings.append(f'Строка {row_idx}: неверный формат срока полезного использования "{row[col_map["useful_life"]]}"')

            cost = None
            if col_map['cost'] != -1 and row[col_map['cost']]:
                try:
                    cost = Decimal(str(row[col_map['cost']]).replace(',', '.'))
                except:
                    warnings.append(f'Строка {row_idx}: неверный формат стоимости "{row[col_map["cost"]]}"')

            serial_number = str(row[col_map['serial']]).strip() if col_map['serial'] != -1 and row[col_map['serial']] else ''
            model = str(row[col_map['model']]).strip() if col_map['model'] != -1 and row[col_map['model']] else ''
            manufacturer = str(row[col_map['manufacturer']]).strip() if col_map['manufacturer'] != -1 and row[col_map['manufacturer']] else ''
            location = str(row[col_map['location']]).strip() if col_map['location'] != -1 and row[col_map['location']] else ''
            status = 'in_use'
            if col_map['status'] != -1 and row[col_map['status']]:
                status_text = str(row[col_map['status']]).strip()
                found = False
                for code, name_choice in Asset.STATUS_CHOICES:
                    if name_choice == status_text:
                        status = code
                        found = True
                        break
                if not found:
                    warnings.append(f'Строка {row_idx}: неизвестный статус "{status_text}", установлен "in_use"')
            notes = str(row[col_map['notes']]).strip() if col_map['notes'] != -1 and row[col_map['notes']] else ''

            # Если есть ошибки, пропускаем запись
            if errors and any(e.startswith(f'Строка {row_idx}:') for e in errors):
                continue

            # Создаём новый объект (импорт только новых, без обновления)
            try:
                Asset.objects.create(
                    inventory_number=inventory_number,
                    name=name,
                    category=category,
                    description='Импортировано из Excel',
                    serial_number=serial_number,
                    manufacturer=manufacturer,
                    model=model,
                    purchase_date=purchase_date,
                    cost=cost,
                    useful_life_months=useful_life_months,
                    location=location,
                    responsible_person=responsible,
                    status=status,
                    notes=notes,
                    imported_from_excel=True,
                )
                created_count += 1
            except Exception as e:
                errors.append(f'Строка {row_idx}: ошибка сохранения: {str(e)}')

        # Формируем итоговое сообщение
        if errors:
            messages.error(request, f'Импорт завершён с ошибками ({len(errors)}). Создано: {created_count}, обновлено: {updated_count}.')
            # Выводим первые 10 ошибок в сообщения
            for err in errors[:10]:
                messages.error(request, err)
            if len(errors) > 10:
                messages.warning(request, f'... и ещё {len(errors)-10} ошибок (смотрите логи)')
        else:
            messages.success(request, f'Импорт успешно завершён. Создано: {created_count}, обновлено: {updated_count}.')

        if warnings:
            messages.warning(request, f'Предупреждения ({len(warnings)}):')
            for warn in warnings[:5]:
                messages.warning(request, warn)

        return redirect('assets:asset_list')

    return render(request, 'assets/import_export.html')


# ---------- УПРАВЛЕНИЕ ФОТО (одиночное) ----------
@login_required
def delete_asset_photo(request, pk):
    """Удаление основного фото объекта"""
    asset = get_object_or_404(Asset, pk=pk)
    if not can_edit_asset(request.user, asset):
        messages.error(request, 'У вас нет прав на удаление фото.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if asset.photo:
        asset.photo.delete(save=False)
        asset.photo = None
        asset.save()
        messages.success(request, 'Основное фото удалено.')
    else:
        messages.warning(request, 'Основное фото отсутствует.')
    return redirect('assets:asset_detail', pk=asset.pk)


@login_required
def upload_asset_photo(request, pk):
    """Загрузка основного фото через отдельную форму (не через редактирование)"""
    asset = get_object_or_404(Asset, pk=pk)
    if not can_edit_asset(request.user, asset):
        messages.error(request, 'У вас нет прав на изменение фото.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if request.method == 'POST' and request.FILES.get('photo'):
        try:
            if asset.photo:
                asset.photo.delete(save=False)
            asset.photo = request.FILES['photo']
            asset.save()
            messages.success(request, 'Основное фото загружено.')
        except Exception as e:
            messages.error(request, f'Ошибка при загрузке фото: {e}')
    else:
        messages.warning(request, 'Файл не выбран или метод не разрешён.')
    return redirect('assets:asset_detail', pk=asset.pk)


# ---------- ГАЛЕРЕЯ (несколько фото) ----------
@login_required
def upload_asset_gallery(request, pk):
    """Загрузка нескольких фото в галерею"""
    asset = get_object_or_404(Asset, pk=pk)
    if not can_edit_asset(request.user, asset):
        messages.error(request, 'У вас нет прав на изменение галереи.')
        return redirect('assets:asset_detail', pk=asset.pk)

    if request.method == 'POST' and request.FILES.getlist('photos'):
        photos = request.FILES.getlist('photos')
        for idx, photo_file in enumerate(photos):
            AssetPhoto.objects.create(
                asset=asset,
                image=photo_file,
                order=idx,
                caption=f'Фото {idx+1}'
            )
        messages.success(request, f'Загружено {len(photos)} фотографий.')
    else:
        messages.warning(request, 'Файлы не выбраны или метод не разрешён.')
    return redirect('assets:asset_detail', pk=asset.pk)


@login_required
def delete_asset_photo_gallery(request, pk, photo_id):
    """Удаление отдельного фото из галереи"""
    asset = get_object_or_404(Asset, pk=pk)
    if not can_edit_asset(request.user, asset):
        messages.error(request, 'Нет прав.')
        return redirect('assets:asset_detail', pk=asset.pk)
    photo = get_object_or_404(AssetPhoto, pk=photo_id, asset=asset)
    photo.delete()  # сигнал удалит файл
    messages.success(request, 'Фото из галереи удалено.')
    return redirect('assets:asset_detail', pk=asset.pk)